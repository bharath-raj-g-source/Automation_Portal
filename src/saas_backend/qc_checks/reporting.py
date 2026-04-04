import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from .common import GREEN_FILL, RED_FILL

def color_excel(output_path, df):
    wb = load_workbook(output_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {name: idx+1 for idx, name in enumerate(headers)}
    qc_columns = [col for col in df.columns if col.endswith("_OK")]

    for col_name in qc_columns:
        if col_name in col_map:
            col_idx = col_map[col_name]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                if val in [True, "True"]:
                    cell.fill = GREEN_FILL
                elif val in [False, "False"]:
                    cell.fill = RED_FILL
    wb.save(output_path)

def generate_summary_sheet(output_path, df):
    wb = load_workbook(output_path)
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")
    qc_columns = [
        col for col in df.columns
        if col.endswith("_OK") or col.endswith("_result")
    ]
    summary_rows = []
    for col in qc_columns:
        series = df[col].astype(str)
        passed = series.eq("True").sum()
        failed = series.eq("False").sum()
        na = series.eq("NA").sum()
        total = passed + failed + na
        summary_rows.append([
            col,
            int(total),
            int(passed),
            int(failed),
            int(na)
        ])
    summary_df = pd.DataFrame(
        summary_rows,
        columns=["Check", "Total Evaluated", "Passed", "Failed", "NA"]
    )
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)
    wb.save(output_path)