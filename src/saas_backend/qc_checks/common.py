import pandas as pd
import numpy as np
import re
import datetime
import logging
from openpyxl.styles import PatternFill

# ----------------------------- Constants -----------------------------
DATE_FORMAT = "%Y-%m-%d"

# Excel color styles
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

# ----------------------------- Helpers -----------------------------
def _find_column(df, candidates):
    if df is None:
        return None
    if not isinstance(candidates, list):
        candidates = [candidates]
    cols_lower = {str(c).lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand is None:
            continue
        k = str(cand).lower().strip()
        if k in cols_lower:
            return cols_lower[k]
    return None

def _is_present(val):
    if val is None:
        return False
    try:
        if pd.isna(val):
            return False
    except Exception:
        pass
    if isinstance(val, (int, float)) and not (isinstance(val, float) and pd.isna(val)):
        return True
    s = str(val).strip()
    if s == "":
        return False
    if s.lower() in ("nan", "none", "-"):
        return False
    return True

def parse_duration_to_minutes(duration_series):
    results = []
    for item in duration_series:
        if pd.isna(item):
            results.append(np.nan)
            continue
        if isinstance(item, (int, float)):
            results.append(float(item))
            continue
        s = str(item).strip()
        try:
            num = float(s)
            results.append(num)
            continue
        except Exception:
            pass
        parts = s.split(':')
        if len(parts) >= 2:
            try:
                hours = float(re.sub(r"[^0-9.]", "", parts[0])) if parts[0] else 0.0
                minutes = float(re.sub(r"[^0-9.]", "", parts[1])) if parts[1] else 0.0
                seconds = 0.0
                if len(parts) >= 3:
                    seconds = float(re.sub(r"[^0-9.]", "", parts[2])) if parts[2] else 0.0
                total_minutes = (hours * 60) + minutes + (seconds / 60)
                results.append(total_minutes)
            except Exception:
                results.append(np.nan)
        else:
            results.append(np.nan)
    return pd.Series(results, index=duration_series.index)

def to_time_str(val):
    """Convert Excel time/float/time/string to HH:MM:SS string."""
    if pd.isna(val):
        return None
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M:%S")
    try:
        if isinstance(val, float) or isinstance(val, int):
            total_seconds = int(val * 24 * 3600)
            h = total_seconds // 3600
            m = (total_seconds % 3600) // 60
            s = total_seconds % 60
            return f"{h:02}:{m:02}:{s:02}"
    except:
        pass
    try:
        t = pd.to_datetime(str(val), errors="coerce")
        if isinstance(t, pd.Timestamp):
            return t.strftime("%H:%M:%S")
    except:
        return None
    return None

def to_date_str(val):
    """Convert Excel date/datetime/string to YYYY-MM-DD."""
    if pd.isna(val):
        return None
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    try:
        d = pd.to_datetime(val, errors="coerce")
        if isinstance(d, pd.Timestamp):
            return d.strftime("%Y-%m-%d")
    except:
        return None
    return None

def combine_parse(date_val, time_val):
    d = to_date_str(date_val)
    t = to_time_str(time_val)
    if not d or not t:
        return pd.NaT
    return pd.to_datetime(f"{d} {t}", errors="coerce")

def auto_sort_bsr(df, bsr_cols):
    """Correct business-level BSR sorting."""
    df = df.copy()
    col_region = _find_column(df, ["Region", "Wider Region"])
    col_country = _find_column(df, ["Country", "Market"])
    col_channel_id = _find_column(df, ["Channel ID"])
    col_date_utc = _find_column(df, ["BSR_UTC_Date", "Date (UTC)", "Date"])
    col_start_utc = _find_column(df, ["Start (UTC)", "Start UTC"])

    if not all([col_country, col_channel_id, col_date_utc, col_start_utc]):
        logging.warning("Auto-sort skipped: required BSR columns missing")
        return df

    df["_utc_dt"] = df.apply(
        lambda r: combine_parse(r[col_date_utc], r[col_start_utc]),
        axis=1
    )

    sort_cols = []
    if col_region:
        sort_cols.append(col_region)
    sort_cols.extend([col_country, col_channel_id, "_utc_dt"])

    df = df.sort_values(
        by=sort_cols,
        ascending=True,
        na_position="last"
    ).reset_index(drop=True)

    df.drop(columns=["_utc_dt"], inplace=True, errors="ignore")
    logging.info("✅ BSR auto-sorted by Region → Country → Channel ID → UTC datetime")
    return df


# -----------------------------------------------------------
# 14️⃣ Home vs Away vs Phase Consistency Check (Updated Logic)
# -----------------------------------------------------------
def home_away_vs_phase_check(df, col_map):
    import re
    import pandas as pd

    result_col = "Home_vs_Away_vs_Phase_OK"
    df[result_col] = "NA"

    # Extract mapping safely
    if isinstance(col_map.get("bsr"), dict):
        b = col_map["bsr"]
    else:
        b = col_map

    def _get_best_col(df, config_key, fallbacks):
        search_terms = []

        config_val = b.get(config_key, [])
        if isinstance(config_val, list):
            search_terms.extend(config_val)
        elif isinstance(config_val, str):
            search_terms.append(config_val)

        search_terms.extend(fallbacks)

        for term in search_terms:
            if not term:
                continue
            if term in df.columns:
                return term
            for actual_col in df.columns:
                if str(actual_col).strip().lower() == str(term).strip().lower():
                    return actual_col
        return None

    # Detect Columns
    col_home = _get_best_col(df, "home_team",
                             ["Home Team", "Home", "Team 1", "Team A", "Home_Team"])

    col_away = _get_best_col(df, "away_team",
                             ["Away Team", "Away", "Team 2", "Team B", "Away_Team"])

    col_phase = _get_best_col(df, "phase_fixture_episode",
                              ["PhaseFixtureEpisode", "Phase", "Fixture",
                               "Combined (translated)", "Event", "Description", "Program"])

    col_indices = {c: i for i, c in enumerate(df.columns)}

    def clean_text(text):
        if pd.isna(text):
            return ""
        return re.sub(r"\s+", " ", str(text).strip().lower())

    # Process Rows
    for idx, row in df.iterrows():

        # Column validation
        if not col_home or not col_phase:
            df.at[idx, result_col] = "Error: Required columns missing"
            continue

        home_team = clean_text(row.get(col_home, ""))
        phase_val = clean_text(row.get(col_phase, ""))

        # If no home team → Not Applicable
        if not home_team:
            df.at[idx, result_col] = "Not Applicable"
            continue

        # Detect Away Team
        away_team = ""

        if col_away:
            away_team = clean_text(row.get(col_away, ""))

        # Fallback positional logic (Home | Vs | Away)
        if not away_team:
            home_idx = col_indices.get(col_home)
            if home_idx is not None and home_idx + 2 < len(df.columns):
                sep_val = clean_text(row.iloc[home_idx + 1])
                if sep_val in ["vs", "v", "vs.", "-", "x", "v."]:
                    away_team = clean_text(row.iloc[home_idx + 2])

        # If either team missing → Not Applicable
        if not home_team or not away_team:
            df.at[idx, result_col] = "Not Applicable"
            continue

        # Final Validation
        if home_team in phase_val and away_team in phase_val:
            df.at[idx, result_col] = True
        else:
            df.at[idx, result_col] = "Home/Away teams do not match PhaseFixtureEpisode"

    return df



def multiple_live_match_check(df, col_map):
    import pandas as pd

    # 1. Initialize result column
    result_col = "Multiple_Live_Match_OK"
    df[result_col] = "NA"

    # Safely get mapping dictionary
    b = col_map.get("bsr", {}) if isinstance(col_map.get("bsr"), dict) else col_map

    # Robust local finder function (self-contained)
    def _get_best_col(df, config_key, fallbacks):
        config_val = b.get(config_key, [])
        search_terms = []
        if isinstance(config_val, list): search_terms.extend(config_val)
        elif isinstance(config_val, str): search_terms.append(config_val)
        search_terms.extend(fallbacks)

        for term in search_terms:
            if not term: continue
            if term in df.columns: return term
            for actual_col in df.columns:
                if str(actual_col).strip().lower() == str(term).strip().lower():
                    return actual_col
        return None

    # 2. Detect Columns (Updated fallbacks based on your actual file)
    col_program_type = _get_best_col(df, "type_of_program", ["Program Type", "Type of Program", "Status", "Live/Repeat"])
    col_market = _get_best_col(df, "market", ["Market", "Region", "Country"])
    col_broadcaster = _get_best_col(df, "broadcaster", ["Broadcaster", "Station"])
    col_channel = _get_best_col(df, "tv_channel", ["Channel", "TV Channel", "Channel ID"])
    col_matchday = _get_best_col(df, "matchday", ["Matchday", "Match Day", "MD"])
    col_match = _get_best_col(df, "phase_fixture_episode", ["PhaseFixtureEpisode", "Phase", "Fixture", "Event", "Combined (translated)"])

    required_cols_map = {
        "Program Type": col_program_type,
        "Market": col_market,
        "Broadcaster": col_broadcaster,
        "Channel": col_channel,
        "Matchday": col_matchday,
        "Match/Phase": col_match
    }

    # 3. Validation: Identify missing headers
    missing_headers = [label for label, val in required_cols_map.items() if not val]
    if missing_headers:
        df[result_col] = f"Error: Missing Headers ({', '.join(missing_headers)})"
        return df

    # 4. Filter for Live rows
    # FIX: Added .str before .lower() to handle the Series correctly
    live_mask = df[col_program_type].astype(str).str.strip().str.lower() == "live"
    
    if not live_mask.any():
        df.loc[:, result_col] = "Not Applicable (No Live Rows)"
        return df

    # 5. Grouping logic to find duplicates
    group_cols = [col_market, col_broadcaster, col_channel, col_matchday, col_match]
    
    # Identify rows that are duplicates within the 'Live' subset
    is_duplicate = df[live_mask].duplicated(subset=group_cols, keep=False)

    # 6. Assign Results
    # Default all Live rows to True
    df.loc[live_mask, result_col] = True
    
    # Update only the flagged duplicates to False
    duplicate_indices = is_duplicate[is_duplicate == True].index
    df.loc[duplicate_indices, result_col] = False

    # Set non-live rows to Not Applicable
    df.loc[~live_mask, result_col] = "Not Applicable"

    return df
