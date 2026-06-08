# import pandas as pd
# import numpy as np
# import re
# import datetime
# import logging
# from openpyxl.styles import PatternFill

# # ----------------------------- Constants -----------------------------
# DATE_FORMAT = "%Y-%m-%d"

# # Excel color styles
# GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
# RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


# # ----------------------------- Helpers -----------------------------
# def _find_column(df, candidates):
#     if df is None:
#         return None
#     if not isinstance(candidates, list):
#         candidates = [candidates]
#     cols_lower = {str(c).lower().strip(): c for c in df.columns}
#     for cand in candidates:
#         if cand is None:
#             continue
#         k = str(cand).lower().strip()
#         if k in cols_lower:
#             return cols_lower[k]
#     return None

# def _is_present(val):
#     if val is None:
#         return False
#     try:
#         if pd.isna(val):
#             return False
#     except Exception:
#         pass
#     if isinstance(val, (int, float)) and not (isinstance(val, float) and pd.isna(val)):
#         return True
#     s = str(val).strip()
#     if s == "":
#         return False
#     if s.lower() in ("nan", "none", "-"):
#         return False
#     return True

# def parse_duration_to_minutes(duration_series):
#     results = []
#     for item in duration_series:
#         if pd.isna(item):
#             results.append(np.nan)
#             continue
#         if isinstance(item, (int, float)):
#             results.append(float(item))
#             continue
#         s = str(item).strip()
#         try:
#             num = float(s)
#             results.append(num)
#             continue
#         except Exception:
#             pass
#         parts = s.split(':')
#         if len(parts) >= 2:
#             try:
#                 hours = float(re.sub(r"[^0-9.]", "", parts[0])) if parts[0] else 0.0
#                 minutes = float(re.sub(r"[^0-9.]", "", parts[1])) if parts[1] else 0.0
#                 seconds = 0.0
#                 if len(parts) >= 3:
#                     seconds = float(re.sub(r"[^0-9.]", "", parts[2])) if parts[2] else 0.0
#                 total_minutes = (hours * 60) + minutes + (seconds / 60)
#                 results.append(total_minutes)
#             except Exception:
#                 results.append(np.nan)
#         else:
#             results.append(np.nan)
#     return pd.Series(results, index=duration_series.index)

# def to_time_str(val):
#     """Convert Excel time/float/time/string to HH:MM:SS string."""
#     if pd.isna(val):
#         return None

#     # if already datetime.time
#     if isinstance(val, datetime.time):
#         return val.strftime("%H:%M:%S")

#     # Excel float time (0.0–1.0)
#     try:
#         if isinstance(val, float) or isinstance(val, int):
#             total_seconds = int(val * 24 * 3600)
#             h = total_seconds // 3600
#             m = (total_seconds % 3600) // 60
#             s = total_seconds % 60
#             return f"{h:02}:{m:02}:{s:02}"
#     except:
#         pass

#     # fallback: string
#     try:
#         t = pd.to_datetime(str(val), errors="coerce")
#         if isinstance(t, pd.Timestamp):
#             return t.strftime("%H:%M:%S")
#     except:
#         return None

#     return None


# def to_date_str(val):
#     """Convert Excel date/datetime/string to YYYY-MM-DD."""
#     if pd.isna(val):
#         return None

#     if isinstance(val, datetime.date):
#         return val.strftime("%Y-%m-%d")

#     try:
#         d = pd.to_datetime(val, errors="coerce")
#         if isinstance(d, pd.Timestamp):
#             return d.strftime("%Y-%m-%d")
#     except:
#         return None

#     return None


# def combine_parse(date_val, time_val):
#     d = to_date_str(date_val)
#     t = to_time_str(time_val)
#     if not d or not t:
#         return pd.NaT
#     return pd.to_datetime(f"{d} {t}", errors="coerce")

# def parse_datetime(date_val, time_val=None):
#     """
#     Robust parser that handles multiple date and time formats.
#     Supports:
#     - DD-MM-YYYY
#     - YYYY-MM-DD
#     - MM/DD/YYYY
#     - Excel serial dates
#     - Datetime objects
#     - Missing or invalid values
#     """
#     try:
#         if pd.isna(date_val):
#             return pd.NaT

#         # Handle Excel serial numbers
#         if isinstance(date_val, (int, float)):
#             date_val = pd.to_datetime(date_val, unit="d", origin="1899-12-30")

#         # Combine date and time if time is provided
#         if time_val is not None and not pd.isna(time_val):
#             datetime_str = f"{date_val} {time_val}"
#         else:
#             datetime_str = date_val

#         return pd.to_datetime(
#             datetime_str,
#             errors="coerce",
#             infer_datetime_format=True,
#             dayfirst=True
#         )

#     except Exception:
#         return pd.NaT


# # ----------------------------- AUTO SORT BSR -----------------------------
# def auto_sort_bsr(df, bsr_cols):
#     df = df.copy()

#     col_region    = _find_column(df, ["Region", "Wider Region"])
#     col_country   = _find_column(df, ["Country", "Market"])
#     col_channel_id = _find_column(df, ["Channel ID"])

#     # ✅ Explicitly avoid matching "Day" as a date column
#     col_date_utc  = None
#     col_start_utc = None
#     for c in df.columns:
#         c_lower = str(c).strip().lower()
#         if c_lower == "day":
#             continue
#         if c_lower in ["date (utc)", "date (utc/gmt)", "bsr_utc_date"]:
#             col_date_utc = c
#         elif c_lower in ["start (utc)", "start utc"]:
#             col_start_utc = c

#     if col_date_utc is None:
#         col_date_utc = _find_column(df, ["BSR_UTC_Date", "Date"])

#     if not all([col_country, col_channel_id, col_date_utc, col_start_utc]):
#         logging.warning("Auto-sort skipped: required BSR columns missing")
#         return df

#     df["_utc_dt"] = df.apply(
#         lambda r: combine_parse(r[col_date_utc], r[col_start_utc]),
#         axis=1
#     )

#     sort_cols = []
#     if col_region:
#         sort_cols.append(col_region)
#     sort_cols.extend([col_country, col_channel_id, "_utc_dt"])

#     df = df.sort_values(
#         by=sort_cols, ascending=True, na_position="last"
#     ).reset_index(drop=True)

#     df.drop(columns=["_utc_dt"], inplace=True, errors="ignore")
#     logging.info("✅ BSR auto-sorted by Region → Country → Channel ID → UTC datetime")
#     return df

# # -----------------------------------------------------------
# # 14️⃣ Home vs Away vs Phase Consistency Check (Updated Logic)
# # -----------------------------------------------------------
# def home_away_vs_phase_check(df, col_map):

#     result_col = "Home_vs_Away_vs_Phase_OK"
#     df[result_col] = "NA"

#     # Extract mapping safely
#     if isinstance(col_map.get("bsr"), dict):
#         b = col_map["bsr"]
#     else:
#         b = col_map

#     def _get_best_col(df, config_key, fallbacks):
#         search_terms = []

#         config_val = b.get(config_key, [])
#         if isinstance(config_val, list):
#             search_terms.extend(config_val)
#         elif isinstance(config_val, str):
#             search_terms.append(config_val)

#         search_terms.extend(fallbacks)

#         for term in search_terms:
#             if not term:
#                 continue
#             if term in df.columns:
#                 return term
#             for actual_col in df.columns:
#                 if str(actual_col).strip().lower() == str(term).strip().lower():
#                     return actual_col
#         return None

#     # Detect Columns
#     col_home = _get_best_col(df, "home_team",
#                              ["Home Team", "Home", "Team 1", "Team A", "Home_Team"])

#     col_away = _get_best_col(df, "away_team",
#                              ["Away Team", "Away", "Team 2", "Team B", "Away_Team"])

#     col_phase = _get_best_col(df, "phase_fixture_episode",
#                               ["PhaseFixtureEpisode", "Phase", "Fixture",
#                                "Combined (translated)", "Event", "Description", "Program"])

#     col_indices = {c: i for i, c in enumerate(df.columns)}

#     def clean_text(text):
#         if pd.isna(text):
#             return ""
#         return re.sub(r"\s+", " ", str(text).strip().lower())
    
#     def normalize_team(team):
#         team = re.sub(r"\(.*?\)","", str(team))
#         return clean_text(team)

#     # Process Rows
#     for idx, row in df.iterrows():

#         # Column validation
#         if not col_home or not col_phase:
#             df.at[idx, result_col] = "Error: Required columns missing"
#             continue

#         home_team = clean_text(row.get(col_home, ""))
#         phase_val = clean_text(row.get(col_phase, ""))

#         # If no home team → Not Applicable
#         if not home_team:
#             df.at[idx, result_col] = "Not Applicable"
#             continue

#         # Detect Away Team
#         away_team = ""

#         if col_away:
#             away_team = clean_text(row.get(col_away, ""))

#         # Fallback positional logic (Home | Vs | Away)
#         if not away_team:
#             home_idx = col_indices.get(col_home)
#             if home_idx is not None and home_idx + 2 < len(df.columns):
#                 sep_val = clean_text(row.iloc[home_idx + 1])
#                 if sep_val in ["vs", "v", "vs.", "-", "x", "v."]:
#                     away_team = clean_text(row.iloc[home_idx + 2])

#         # If either team missing → Not Applicable
#         if not home_team or not away_team:
#             df.at[idx, result_col] = "Not Applicable"
#             continue
#         home_team_n = normalize_team(home_team)
#         away_team_n = normalize_team(away_team)
#         phase_val_n = normalize_team(phase_val)

#         # Final Validation
#         if home_team_n in phase_val_n and away_team_n in phase_val_n:
#             df.at[idx, result_col] = "OK"
#         else:
#             df.at[idx, result_col] = "Home/Away teams do not match PhaseFixtureEpisode"

#     return df



# def multiple_live_match_check(df, col_map):

#     result_col = "Multiple_Live_Match_OK"

#     df = df.copy()

#     # Default value
#     df[result_col] = "NA"

#     # -------------------------------------------------------
#     # Safe config mapping
#     # -------------------------------------------------------
#     b = col_map.get("bsr", {}) if isinstance(col_map.get("bsr"), dict) else col_map

#     # -------------------------------------------------------
#     # Helper: find best matching column
#     # -------------------------------------------------------
#     def _get_best_col(df, config_key, fallbacks):

#         config_val = b.get(config_key, [])
#         search_terms = []

#         if isinstance(config_val, list):
#             search_terms.extend(config_val)

#         elif isinstance(config_val, str):
#             search_terms.append(config_val)

#         search_terms.extend(fallbacks)

#         for term in search_terms:

#             if not term:
#                 continue

#             # Exact match
#             if term in df.columns:
#                 return term

#             # Case-insensitive match
#             for actual_col in df.columns:

#                 if (
#                     str(actual_col).strip().lower()
#                     == str(term).strip().lower()
#                 ):
#                     return actual_col

#         return None

#     # -------------------------------------------------------
#     # Detect columns
#     # -------------------------------------------------------

#     col_program_type = _get_best_col(
#         df,
#         "type_of_program",
#         ["Program Type", "Type of Program", "Status", "Live/Repeat"]
#     )

#     col_market = _get_best_col(
#         df,
#         "market",
#         ["Market", "Region", "Country"]
#     )

#     col_broadcaster = _get_best_col(
#         df,
#         "broadcaster",
#         ["Broadcaster", "Station"]
#     )

#     col_channel = _get_best_col(
#         df,
#         "tv_channel",
#         ["Channel", "TV Channel", "Channel ID"]
#     )

#     col_date = _get_best_col(
#         df,
#         "date",
#         ["Date", "BSR_Local_Date", "BSR Date"]
#     )

#     col_start = _get_best_col(
#         df,
#         "start_time",
#         ["Start", "Program Start", "Start Time"]
#     )

#     col_end = _get_best_col(
#         df,
#         "end_time",
#         ["End", "End Time"]
#     )

#     # IMPORTANT:
#     # Use the strongest match identifier possible
#     col_match = _get_best_col(
#         df,
#         "phase_fixture_episode",
#         [
#             "PhaseFixtureEpisode",
#             "Program title",
#             "Combined (translated)"
#         ]
#     )

#     # -------------------------------------------------------
#     # Validate required columns
#     # -------------------------------------------------------

#     required_cols_map = {
#         "Program Type": col_program_type,
#         "Market": col_market,
#         "Broadcaster": col_broadcaster,
#         "Channel": col_channel,
#         "Date": col_date,
#         "Start": col_start,
#         "End": col_end,
#         "Match Identifier": col_match
#     }

#     missing_headers = [
#         label
#         for label, val in required_cols_map.items()
#         if not val
#     ]

#     if missing_headers:

#         df[result_col] = (
#             f"Error: Missing Headers ({', '.join(missing_headers)})"
#         )

#         return df

#     # -------------------------------------------------------
#     # LIVE rows only
#     # -------------------------------------------------------

#     live_mask = (
#         df[col_program_type]
#         .astype(str)
#         .str.strip()
#         .str.lower()
#         .eq("live")
#     )

#     # No live rows
#     if not live_mask.any():

#         df[result_col] = "Not Applicable"

#         return df

#     # -------------------------------------------------------
#     # Normalize fields
#     # -------------------------------------------------------

#     for col in [
#         col_market,
#         col_broadcaster,
#         col_channel,
#         col_date,
#         col_start,
#         col_end,
#         col_match
#     ]:

#         df[col] = (
#             df[col]
#             .astype(str)
#             .str.strip()
#             .str.lower()
#         )

#     # -------------------------------------------------------
#     # Duplicate detection logic
#     # -------------------------------------------------------

#     # Exact LIVE duplicate definition:
#     # same market + broadcaster + channel
#     # same date + start + end
#     # same match/program

#     group_cols = [
#         col_market,
#         col_broadcaster,
#         col_channel,
#         col_date,
#         col_start,
#         col_end,
#         col_match
#     ]

#     live_df = df[live_mask]

#     duplicate_mask = live_df.duplicated(
#         subset=group_cols,
#         keep="first"
#     )

#     # -------------------------------------------------------
#     # Assign results
#     # -------------------------------------------------------

#     # Default for live rows
#     df.loc[live_mask, result_col] = "True"

#     # Only repeated duplicate rows become False
#     df.loc[
#         duplicate_mask[duplicate_mask].index,
#         result_col
#     ] = "False"

#     # Non-live rows
#     df.loc[~live_mask, result_col] = "Not Applicable"

#     return df



import pandas as pd
import numpy as np
import re
import datetime
import logging
from openpyxl.styles import PatternFill



# ----------------------------- Constants -----------------------------
DATE_FORMAT = "%Y-%m-%d"


# # Excel color styles
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
# ----------------------------- Helpers -----------------------------
def _find_column(df, candidates):
    if df is None:
        return None
    if not isinstance(candidates, list):
        candidates = [candidates]
    cols_lower = {str(c).lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand is None:
            continue
        k = str(cand).lower().strip()
        if k in cols_lower:
            return cols_lower[k]
    return None

def _is_present(val):
    if val is None:
        return False
    try:
        if pd.isna(val):
            return False
    except Exception:
        pass
    if isinstance(val, (int, float)) and not (isinstance(val, float) and pd.isna(val)):
        return True
    s = str(val).strip()
    if s == "":
        return False
    if s.lower() in ("nan", "none", "-"):
        return False
    return True

def parse_duration_to_minutes(duration_series):
    results = []
    for item in duration_series:
        if pd.isna(item):
            results.append(np.nan)
            continue
        if isinstance(item, (int, float)):
            results.append(float(item))
            continue
        s = str(item).strip()
        try:
            num = float(s)
            results.append(num)
            continue
        except Exception:
            pass
        parts = s.split(':')
        if len(parts) >= 2:
            try:
                hours = float(re.sub(r"[^0-9.]", "", parts[0])) if parts[0] else 0.0
                minutes = float(re.sub(r"[^0-9.]", "", parts[1])) if parts[1] else 0.0
                seconds = 0.0
                if len(parts) >= 3:
                    seconds = float(re.sub(r"[^0-9.]", "", parts[2])) if parts[2] else 0.0
                total_minutes = (hours * 60) + minutes + (seconds / 60)
                results.append(total_minutes)
            except Exception:
                results.append(np.nan)
        else:
            results.append(np.nan)
    return pd.Series(results, index=duration_series.index)

def to_time_str(val):
    """Convert Excel time/float/time/string to HH:MM:SS string."""
    if pd.isna(val):
        return None

    # if already datetime.time
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M:%S")

    # Excel float time (0.0–1.0)
    try:
        if isinstance(val, float) or isinstance(val, int):
            total_seconds = int(val * 24 * 3600)
            h = total_seconds // 3600
            m = (total_seconds % 3600) // 60
            s = total_seconds % 60
            return f"{h:02}:{m:02}:{s:02}"
    except:
        pass

    # fallback: string
    try:
        t = pd.to_datetime(str(val), errors="coerce")
        if isinstance(t, pd.Timestamp):
            return t.strftime("%H:%M:%S")
    except:
        return None

    return None


def to_date_str(val):
    """Convert Excel date/datetime/string to YYYY-MM-DD."""
    if pd.isna(val):
        return None

    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")

    try:
        d = pd.to_datetime(val, errors="coerce")
        if isinstance(d, pd.Timestamp):
            return d.strftime("%Y-%m-%d")
    except:
        return None

    return None


def combine_parse(date_val, time_val):
    d = to_date_str(date_val)
    t = to_time_str(time_val)
    if not d or not t:
        return pd.NaT
    return pd.to_datetime(f"{d} {t}", errors="coerce")

def parse_datetime(date_val, time_val=None):
    """
    Robust parser that handles multiple date and time formats.
    Supports:
    - DD-MM-YYYY
    - YYYY-MM-DD
    - MM/DD/YYYY
    - Excel serial dates
    - Datetime objects
    - Missing or invalid values
    """
    try:
        if pd.isna(date_val):
            return pd.NaT

        # Handle Excel serial numbers
        if isinstance(date_val, (int, float)):
            date_val = pd.to_datetime(date_val, unit="d", origin="1899-12-30")

        # Combine date and time if time is provided
        if time_val is not None and not pd.isna(time_val):
            datetime_str = f"{date_val} {time_val}"
        else:
            datetime_str = date_val

        return pd.to_datetime(
            datetime_str,
            errors="coerce",
            infer_datetime_format=True,
            dayfirst=True
        )

    except Exception:
        return pd.NaT


# ----------------------------- AUTO SORT BSR -----------------------------
def auto_sort_bsr(df, bsr_cols):
    df = df.copy()

    col_region    = _find_column(df, ["Region", "Wider Region"])
    col_country   = _find_column(df, ["Country", "Market"])
    col_channel_id = _find_column(df, ["Channel ID"])

    # ✅ Explicitly avoid matching "Day" as a date column
    col_date_utc  = None
    col_start_utc = None
    for c in df.columns:
        c_lower = str(c).strip().lower()
        if c_lower == "day":
            continue
        if c_lower in ["date (utc)", "date (utc/gmt)", "bsr_utc_date"]:
            col_date_utc = c
        elif c_lower in ["start (utc)", "start utc"]:
            col_start_utc = c

    if col_date_utc is None:
        col_date_utc = _find_column(df, ["BSR_UTC_Date", "Date"])

    if not all([col_country, col_channel_id, col_date_utc, col_start_utc]):
        logging.warning("Auto-sort skipped: required BSR columns missing")
        return df

    df["_utc_dt"] = df.apply(
        lambda r: combine_parse(r[col_date_utc], r[col_start_utc]),
        axis=1
    )

    sort_cols = []
    if col_region:
        sort_cols.append(col_region)
    sort_cols.extend([col_country, col_channel_id, "_utc_dt"])

    df = df.sort_values(
        by=sort_cols, ascending=True, na_position="last"
    ).reset_index(drop=True)

    df.drop(columns=["_utc_dt"], inplace=True, errors="ignore")
    logging.info("✅ BSR auto-sorted by Region → Country → Channel ID → UTC datetime")
    return df

# -----------------------------------------------------------
# 14️⃣ Home vs Away vs Phase Consistency Check (Updated Logic)
# -----------------------------------------------------------
def home_away_vs_phase_check(df, col_map):

    result_col = "Home_vs_Away_vs_Phase_OK"
    df[result_col] = "NA"

    # Extract mapping safely
    if isinstance(col_map.get("bsr"), dict):
        b = col_map["bsr"]
    else:
        b = col_map

    def _get_best_col(df, config_key, fallbacks):
        search_terms = []

        config_val = b.get(config_key, [])
        if isinstance(config_val, list):
            search_terms.extend(config_val)
        elif isinstance(config_val, str):
            search_terms.append(config_val)

        search_terms.extend(fallbacks)

        for term in search_terms:
            if not term:
                continue
            if term in df.columns:
                return term
            for actual_col in df.columns:
                if str(actual_col).strip().lower() == str(term).strip().lower():
                    return actual_col
        return None

    # Detect Columns
    col_home = _get_best_col(df, "home_team",
                             ["Home Team", "Home", "Team 1", "Team A", "Home_Team"])

    col_away = _get_best_col(df, "away_team",
                             ["Away Team", "Away", "Team 2", "Team B", "Away_Team"])

    col_phase = _get_best_col(df, "phase_fixture_episode",
                              ["PhaseFixtureEpisode", "Phase", "Fixture",
                               "Combined (translated)", "Event", "Description", "Program"])

    col_indices = {c: i for i, c in enumerate(df.columns)}

    def clean_text(text):
        if pd.isna(text):
            return ""
        return re.sub(r"\s+", " ", str(text).strip().lower())
    
    def normalize_team(team):
        team = re.sub(r"\(.*?\)","", str(team))
        return clean_text(team)

    # Process Rows
    for idx, row in df.iterrows():

        # Column validation
        if not col_home or not col_phase:
            df.at[idx, result_col] = "Error: Required columns missing"
            continue

        home_team = clean_text(row.get(col_home, ""))
        phase_val = clean_text(row.get(col_phase, ""))

        # If no home team → Not Applicable
        if not home_team:
            df.at[idx, result_col] = "Not Applicable"
            continue

        # Detect Away Team
        away_team = ""

        if col_away:
            away_team = clean_text(row.get(col_away, ""))

        # Fallback positional logic (Home | Vs | Away)
        if not away_team:
            home_idx = col_indices.get(col_home)
            if home_idx is not None and home_idx + 2 < len(df.columns):
                sep_val = clean_text(row.iloc[home_idx + 1])
                if sep_val in ["vs", "v", "vs.", "-", "x", "v."]:
                    away_team = clean_text(row.iloc[home_idx + 2])

        # If either team missing → Not Applicable
        if not home_team or not away_team:
            df.at[idx, result_col] = "Not Applicable"
            continue
        home_team_n = normalize_team(home_team)
        away_team_n = normalize_team(away_team)
        phase_val_n = normalize_team(phase_val)

        # Final Validation
        if home_team_n in phase_val_n and away_team_n in phase_val_n:
            df.at[idx, result_col] = "OK"
        else:
            df.at[idx, result_col] = "Home/Away teams do not match PhaseFixtureEpisode"

    return df



def multiple_live_match_check(df, col_map):

    result_col = "Multiple_Live_Match_OK"

    df = df.copy()

    # Default value
    df[result_col] = "NA"

    # -------------------------------------------------------
    # Safe config mapping
    # -------------------------------------------------------
    b = col_map.get("bsr", {}) if isinstance(col_map.get("bsr"), dict) else col_map

    # -------------------------------------------------------
    # Helper: find best matching column
    # -------------------------------------------------------
    def _get_best_col(df, config_key, fallbacks):

        config_val = b.get(config_key, [])
        search_terms = []

        if isinstance(config_val, list):
            search_terms.extend(config_val)

        elif isinstance(config_val, str):
            search_terms.append(config_val)

        search_terms.extend(fallbacks)

        for term in search_terms:

            if not term:
                continue

            # Exact match
            if term in df.columns:
                return term

            # Case-insensitive match
            for actual_col in df.columns:

                if (
                    str(actual_col).strip().lower()
                    == str(term).strip().lower()
                ):
                    return actual_col

        return None

    # -------------------------------------------------------
    # Detect columns
    # -------------------------------------------------------

    col_program_type = _get_best_col(
        df,
        "type_of_program",
        ["Program Type", "Type of Program", "Status", "Live/Repeat"]
    )

    col_market = _get_best_col(
        df,
        "market",
        ["Market", "Region", "Country"]
    )

    col_broadcaster = _get_best_col(
        df,
        "broadcaster",
        ["Broadcaster", "Station"]
    )

    col_channel = _get_best_col(
        df,
        "tv_channel",
        ["Channel", "TV Channel", "Channel ID"]
    )

    col_date = _get_best_col(
        df,
        "date",
        ["Date", "BSR_Local_Date", "BSR Date"]
    )

    col_start = _get_best_col(
        df,
        "start_time",
        ["Start", "Program Start", "Start Time"]
    )

    col_end = _get_best_col(
        df,
        "end_time",
        ["End", "End Time"]
    )

    # IMPORTANT:
    # Use the strongest match identifier possible
    col_match = _get_best_col(
        df,
        "phase_fixture_episode",
        [
            "PhaseFixtureEpisode",
            "Program title",
            "Combined (translated)"
        ]
    )

    # -------------------------------------------------------
    # Validate required columns
    # -------------------------------------------------------

    required_cols_map = {
        "Program Type": col_program_type,
        "Market": col_market,
        "Broadcaster": col_broadcaster,
        "Channel": col_channel,
        "Date": col_date,
        "Start": col_start,
        "End": col_end,
        "Match Identifier": col_match
    }

    missing_headers = [
        label
        for label, val in required_cols_map.items()
        if not val
    ]

    if missing_headers:

        df[result_col] = (
            f"Error: Missing Headers ({', '.join(missing_headers)})"
        )

        return df

    # -------------------------------------------------------
    # LIVE rows only
    # -------------------------------------------------------

    live_mask = (
        df[col_program_type]
        .astype(str)
        .str.strip()
        .str.lower()
        .eq("live")
    )

    # No live rows
    if not live_mask.any():

        df[result_col] = "Not Applicable"

        return df

    # -------------------------------------------------------
    # Normalize fields
    # -------------------------------------------------------

    for col in [
        col_market,
        col_broadcaster,
        col_channel,
        col_date,
        col_start,
        col_end,
        col_match
    ]:

        df[col] = (
            df[col]
            .astype(str)
            .str.strip()
            .str.lower()
        )

    # -------------------------------------------------------
    # Duplicate detection logic
    # -------------------------------------------------------

    # Exact LIVE duplicate definition:
    # same market + broadcaster + channel
    # same date + start + end
    # same match/program

    group_cols = [
        col_market,
        col_broadcaster,
        col_channel,
        col_date,
        col_start,
        col_end,
        col_match
    ]

    live_df = df[live_mask]

    duplicate_mask = live_df.duplicated(
        subset=group_cols,
        keep="first"
    )

    # -------------------------------------------------------
    # Assign results
    # -------------------------------------------------------

    # Default for live rows
    df.loc[live_mask, result_col] = "True"

    # Only repeated duplicate rows become False
    df.loc[
        duplicate_mask[duplicate_mask].index,
        result_col
    ] = "False"

    # Non-live rows
    df.loc[~live_mask, result_col] = "Not Applicable"

    return df
