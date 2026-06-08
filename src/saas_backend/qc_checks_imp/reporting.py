import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from .common import GREEN_FILL, RED_FILL

def color_excel(output_path, df):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    wb = load_workbook(output_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {name: idx + 1 for idx, name in enumerate(headers)}

    # ✅ Support both old and new QC result columns
    qc_columns = [
        col for col in df.columns
        if col.endswith("_OK") or col.endswith("_result")
    ]

    for col_name in qc_columns:
        if col_name in col_map:
            col_idx = col_map[col_name]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                if val in [True, "True"]:
                    cell.fill = GREEN_FILL
                elif val in [False, "False"]:
                    cell.fill = RED_FILL
                elif val in ["NA", None]:
                    cell.fill = GREY_FILL

    wb.save(output_path)
# -----------------------------------------------------------
# Summary Sheet
def generate_summary_sheet(output_path, df):
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = load_workbook(output_path)

    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws = wb.create_sheet("Summary")

    # ✅ Support both old and new QC result columns
    qc_columns = [
        col for col in df.columns
        if col.endswith("_OK") or col.endswith("_result")
    ]

    summary_rows = []

    for col in qc_columns:
        series = df[col].astype(str)

        passed = series.eq("True").sum()
        failed = series.eq("False").sum()
        na = series.eq("NA").sum()
        total = passed + failed + na

        summary_rows.append([
            col,
            int(total),
            int(passed),
            int(failed),
            int(na)
        ])

    summary_df = pd.DataFrame(
        summary_rows,
        columns=["Check", "Total Evaluated", "Passed", "Failed", "NA"]
    )

    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)

    wb.save(output_path)


import pandas as pd


def build_summary_dataframe(df):

    qc_columns = [
        col for col in df.columns
        if col.endswith("_OK")
        or col.endswith("_result")
    ]

    summary_rows = []

    for col in qc_columns:

        series = (
            df[col]
            .astype(str)
            .str.upper()
        )

        passed = series.eq("TRUE").sum()

        failed = series.isin(
            ["FALSE", "FAILED", "0"]
        ).sum()

        na_count = series.isin(
            ["NA", "NAN", ""]
        ).sum()

        total = passed + failed + na_count

        summary_rows.append({
            "Check": col,
            "Total Evaluated": int(total),
            "Passed": int(passed),
            "Failed": int(failed),
            "NA": int(na_count)
        })

    return pd.DataFrame(summary_rows)


def export_gqc_report(
    output_path,
    qc_df,
    fixtures_df=None
):

    summary_df = build_summary_dataframe(qc_df)

    with pd.ExcelWriter(
        output_path,
        engine="xlsxwriter"
    ) as writer:

        qc_df.to_excel(
            writer,
            sheet_name="QC Results",
            index=False
        )

        summary_df.to_excel(
            writer,
            sheet_name="Summary",
            index=False
        )

        if (
            fixtures_df is not None
            and not fixtures_df.empty
        ):

            fixtures_df.to_excel(
                writer,
                sheet_name="Original Fixtures",
                index=False
            )