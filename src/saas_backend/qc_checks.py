import pandas as pd
import re
import math
import datetime
import os
import numpy as np
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

DATE_FORMAT = "%Y-%m-%d"

# Excel color styles
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

# ----------------------------- Helpers -----------------------------
def _find_column(df, candidates):
    if df is None:
        return None
    if not isinstance(candidates, list):
        candidates = [candidates]
    cols_lower = {str(c).lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand is None:
            continue
        k = str(cand).lower().strip()
        if k in cols_lower:
            return cols_lower[k]
    return None

def _is_present(val):
    if val is None:
        return False
    try:
        if pd.isna(val):
            return False
    except Exception:
        pass
    if isinstance(val, (int, float)) and not (isinstance(val, float) and pd.isna(val)):
        return True
    s = str(val).strip()
    if s == "":
        return False
    if s.lower() in ("nan", "none", "-"):
        return False
    return True



def parse_duration_to_minutes(duration_series):
    results = []
    for item in duration_series:
        if pd.isna(item):
            results.append(np.nan)
            continue
        if isinstance(item, (int, float)):
            results.append(float(item))
            continue
        s = str(item).strip()
        try:
            num = float(s)
            results.append(num)
            continue
        except Exception:
            pass
        parts = s.split(':')
        if len(parts) >= 2:
            try:
                hours = float(re.sub(r"[^0-9.]", "", parts[0])) if parts[0] else 0.0
                minutes = float(re.sub(r"[^0-9.]", "", parts[1])) if parts[1] else 0.0
                seconds = 0.0
                if len(parts) >= 3:
                    seconds = float(re.sub(r"[^0-9.]", "", parts[2])) if parts[2] else 0.0
                total_minutes = (hours * 60) + minutes + (seconds / 60)
                results.append(total_minutes)
            except Exception:
                results.append(np.nan)
        else:
            results.append(np.nan)
    return pd.Series(results, index=duration_series.index)

def to_time_str(val):
    """Convert Excel time/float/time/string to HH:MM:SS string."""
    if pd.isna(val):
        return None

    # if already datetime.time
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M:%S")

    # Excel float time (0.0–1.0)
    try:
        if isinstance(val, float) or isinstance(val, int):
            total_seconds = int(val * 24 * 3600)
            h = total_seconds // 3600
            m = (total_seconds % 3600) // 60
            s = total_seconds % 60
            return f"{h:02}:{m:02}:{s:02}"
    except:
        pass

    # fallback: string
    try:
        t = pd.to_datetime(str(val), errors="coerce")
        if isinstance(t, pd.Timestamp):
            return t.strftime("%H:%M:%S")
    except:
        return None

    return None


def to_date_str(val):
    """Convert Excel date/datetime/string to YYYY-MM-DD."""
    if pd.isna(val):
        return None

    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")

    try:
        d = pd.to_datetime(val, errors="coerce")
        if isinstance(d, pd.Timestamp):
            return d.strftime("%Y-%m-%d")
    except:
        return None

    return None


def combine_parse(date_val, time_val):
    d = to_date_str(date_val)
    t = to_time_str(time_val)
    if not d or not t:
        return pd.NaT
    return pd.to_datetime(f"{d} {t}", errors="coerce")

def _normalize_excel_columns(df):
    WEEKDAY_NAMES = {
        "monday","tuesday","wednesday","thursday","friday","saturday","sunday",
        "mon","tue","wed","thu","fri","sat","sun"
    }

    def excel_float_to_time(v):
        """Convert Excel time fraction (including 0 and 1) to HH:MM:SS string."""
        try:
            f = float(v)
            # f==1.0 means exactly 24:00 which is 00:00:00 next day — treat as 00:00:00
            if f >= 1.0:
                f = f - int(f)
            total_seconds = round(f * 86400)
            h = total_seconds // 3600
            m = (total_seconds % 3600) // 60
            s = total_seconds % 60
            return f"{h:02d}:{m:02d}:{s:02d}"
        except Exception:
            return v

    def convert_value(v, target):
        if v is None:
            return v
        # Check for pandas/numpy NA carefully — avoid calling pd.isna on non-scalar
        try:
            if pd.isna(v):
                return v
        except Exception:
            pass

        # datetime.time → time string
        if isinstance(v, datetime.time):
            return v.strftime("%H:%M:%S") if target == "time" else v

        # Full Timestamp or datetime → split by target
        if isinstance(v, (pd.Timestamp, datetime.datetime)):
            if target == "date":
                return v.strftime("%Y-%m-%d")
            elif target == "time":
                return v.strftime("%H:%M:%S")

        # date only (no time component)
        if isinstance(v, datetime.date):
            return v.strftime("%Y-%m-%d") if target == "date" else v

        # ✅ Numeric — MUST handle 0 explicitly (midnight)
        # Use explicit type check, not truthiness, so 0 is not skipped
        if type(v) in (int, float) or (hasattr(np, 'integer') and isinstance(v, np.integer)) \
                or (hasattr(np, 'floating') and isinstance(v, np.floating)):
            numeric = float(v)
            if target == "time":
                return excel_float_to_time(numeric)
            elif target == "date" and numeric > 2:
                try:
                    converted = pd.Timestamp("1899-12-30") + pd.to_timedelta(numeric, unit="D")
                    return converted.strftime("%Y-%m-%d")
                except Exception:
                    return v

        # String fallback
        try:
            s = str(v).strip()
            parsed = pd.to_datetime(s, errors="coerce")
            if pd.notna(parsed):
                if target == "date":
                    return parsed.strftime("%Y-%m-%d")
                elif target == "time":
                    return parsed.strftime("%H:%M:%S")
        except Exception:
            pass

        return v

    for col in df.columns:
        col_lower = str(col).strip().lower()

        if col_lower == "day":
            continue

        is_date_col = "date" in col_lower
        is_time_col = any(kw in col_lower for kw in ["start", "end"])

        if not is_date_col and not is_time_col:
            continue

        # Skip if column contains weekday names
        sample_vals = df[col].dropna().head(20)
        if sample_vals.apply(
            lambda v: str(v).strip().lower() in WEEKDAY_NAMES
        ).any():
            continue

        target = "date" if is_date_col else "time"
        df[col] = df[col].apply(lambda v: convert_value(v, target))

    return df

def parse_datetime(date_val, time_val=None):
    """
    Robust parser that handles multiple date and time formats.
    Supports:
    - DD-MM-YYYY
    - YYYY-MM-DD
    - MM/DD/YYYY
    - Excel serial dates
    - Datetime objects
    - Missing or invalid values
    """
    try:
        if pd.isna(date_val):
            return pd.NaT

        # Handle Excel serial numbers
        if isinstance(date_val, (int, float)):
            date_val = pd.to_datetime(date_val, unit="d", origin="1899-12-30")

        # Combine date and time if time is provided
        if time_val is not None and not pd.isna(time_val):
            datetime_str = f"{date_val} {time_val}"
        else:
            datetime_str = date_val

        return pd.to_datetime(
            datetime_str,
            errors="coerce",
            infer_datetime_format=True,
            dayfirst=True
        )

    except Exception:
        return pd.NaT


# ----------------------------- AUTO SORT BSR -----------------------------
def auto_sort_bsr(df, bsr_cols):
    df = df.copy()

    col_region    = _find_column(df, ["Region", "Wider Region"])
    col_country   = _find_column(df, ["Country", "Market"])
    col_channel_id = _find_column(df, ["Channel ID"])

    # ✅ Explicitly avoid matching "Day" as a date column
    col_date_utc  = None
    col_start_utc = None
    for c in df.columns:
        c_lower = str(c).strip().lower()
        if c_lower == "day":
            continue
        if c_lower in ["date (utc)", "date (utc/gmt)", "bsr_utc_date"]:
            col_date_utc = c
        elif c_lower in ["start (utc)", "start utc"]:
            col_start_utc = c

    if col_date_utc is None:
        col_date_utc = _find_column(df, ["BSR_UTC_Date", "Date"])

    if not all([col_country, col_channel_id, col_date_utc, col_start_utc]):
        logging.warning("Auto-sort skipped: required BSR columns missing")
        return df

    df["_utc_dt"] = df.apply(
        lambda r: combine_parse(r[col_date_utc], r[col_start_utc]),
        axis=1
    )

    sort_cols = []
    if col_region:
        sort_cols.append(col_region)
    sort_cols.extend([col_country, col_channel_id, "_utc_dt"])

    df = df.sort_values(
        by=sort_cols, ascending=True, na_position="last"
    ).reset_index(drop=True)

    df.drop(columns=["_utc_dt"], inplace=True, errors="ignore")
    logging.info("✅ BSR auto-sorted by Region → Country → Channel ID → UTC datetime")
    return df

# ----------------------------- 1️⃣ Detect Monitoring Period -----------------------------
def detect_period_from_rosco(rosco_path):
    df = pd.read_excel(rosco_path, header=None)

    label_col = df.iloc[:, 1].astype(str)
    period_row_mask = label_col.str.contains(
        "monitoring period", case=False, na=False
    )

    if not period_row_mask.any():
        raise ValueError("Missing monitoring period label in Column B of Rosco")

    row_idx = period_row_mask.idxmax()

    if df.shape[1] <= 2:
        raise ValueError(
            f"Missing monitoring period, please fill cell C{row_idx + 1} of Rosco"
        )

    user_input_text = str(df.iloc[row_idx, 2]).strip()

    if not user_input_text or user_input_text.lower() == "nan":
        raise ValueError(
            f"Missing monitoring period in cell C{row_idx + 1} of Rosco"
        )

    found = re.findall(r"\d{4}-\d{2}-\d{2}", user_input_text)

    if len(found) < 2:
        raise ValueError(
            f"Invalid date format in cell C{row_idx + 1}. "
            "Expected two dates (YYYY-MM-DD)."
        )

    start_date = pd.to_datetime(found[0], format=DATE_FORMAT).date()
    end_date   = pd.to_datetime(found[1], format=DATE_FORMAT).date()

    if start_date > end_date:
        raise ValueError(
            f"Invalid monitoring period in cell C{row_idx + 1}: "
            "start date is after end date"
        )

    return start_date, end_date

def parse_frontend_dates(start_date_str, end_date_str):
    if not start_date_str or not end_date_str:
        raise ValueError("Missing monitoring period. Please select both Start and End dates in the UI.")
    
    try:    
        # Convert to Timestamp objects first
        s_dt = pd.to_datetime(start_date_str, errors="coerce")
        e_dt = pd.to_datetime(end_date_str, errors="coerce")

        # Check if they became NaT
        if pd.isna(s_dt) or pd.isna(e_dt):
            raise ValueError("Invalid date format provided.")

        # ONLY call .date() after confirming it's not NaT
        start_date = s_dt.date()
        end_date = e_dt.date()
        
    except Exception as e:
        if "Invalid date format" in str(e): raise
        raise ValueError(f"Error parsing dates: {str(e)}")

    if start_date > end_date:
        raise ValueError("Invalid duration: Start Date cannot be after End Date.")
        
    return start_date, end_date

    
# ----------------------------- 2️⃣ Load BSR -----------------------------
def detect_header_row_in_sheet(bsr_path, sheet_name):
    df_sample = pd.read_excel(
        bsr_path,
        sheet_name=sheet_name,   #  LOCKED to this sheet
        header=None,
        nrows=200
    )

    for i, row in df_sample.iterrows():
        row_str = " ".join(row.dropna().astype(str)).lower()

        if "region" in row_str and "market" in row_str and "broadcaster" in row_str:
            return i

        if "date" in row_str and ("utc" in row_str or "gmt" in row_str):
            return i

    raise ValueError(
        f"Header not found in sheet '{sheet_name}'"
    )


# def df(bsr_path):
#     header_row = detect_header_row(bsr_path)
#     df = pd.read_excel(bsr_path, header=header_row)
#     df.columns = [str(c).strip() for c in df.columns]
#     return df

def load_bsr(bsr_path):
    xl = pd.ExcelFile(bsr_path)

    allowed_sheets = {"worksheet", "database"}
    target_sheet = None

    for sheet in xl.sheet_names:
        if sheet.strip().lower() in allowed_sheets:
            target_sheet = sheet
            break

    if not target_sheet:
        raise ValueError(
            f"No valid sheet ('Worksheet' or 'Database') found in {os.path.basename(bsr_path)}"
        )

    header_row = detect_header_row_in_sheet(bsr_path, target_sheet)

    # ── Step 1: peek at the header to find the Day column index ──
    header_df = pd.read_excel(
        bsr_path,
        sheet_name=target_sheet,
        header=header_row,
        nrows=0          # just the header, no data rows
    )
    header_df.columns = [str(c).strip() for c in header_df.columns]

    # Build dtype override: force every column whose name is exactly "Day" to str
    dtype_overrides = {}
    for col in header_df.columns:
        if col.strip().lower() == "day":
            dtype_overrides[col] = str

    # ── Step 2: load with dtype overrides ──
    df = pd.read_excel(
        bsr_path,
        sheet_name=target_sheet,
        header=header_row,
        dtype=dtype_overrides   # only Day is forced to str; everything else is auto
    )

    df.columns = [str(c).strip() for c in df.columns]

    # ── Step 3: clean up any "nan" strings that dtype=str introduces ──
    for col in df.columns:
        if col.strip().lower() == "day":
            df[col] = df[col].apply(
                lambda v: "" if str(v).strip().lower() in ("nan", "none", "") else str(v).strip()
            )

    # ── Step 4: normalize date/time columns ──
    df = _normalize_excel_columns(df)

    return df

# ----------------------------- 3️⃣ Period Check -----------------------------
def period_check(bsr_df, start_date, end_date):
    bsr_df = bsr_df.copy()

    start_ts = pd.to_datetime(start_date).normalize()
    end_ts   = pd.to_datetime(end_date).normalize()

    utc_col = None
    local_col = None

    for c in bsr_df.columns:
        cname = str(c).lower().replace(" ", "").replace("_", "")
        # ✅ NEVER treat the "Day" column as a date column
        if str(c).strip().lower() == "day":
            continue
        if "date" in cname and "utc" in cname:
            utc_col = c
        elif cname == "date":
            local_col = c

    if utc_col is None and local_col is None:
        raise ValueError("No valid date columns found in BSR")

    def normalize_dt(series):
        if pd.api.types.is_datetime64_any_dtype(series):
            return series.dt.normalize()
        return pd.to_datetime(series, errors="coerce").dt.normalize()

    bsr_df["BSR_UTC_Date"] = (
        normalize_dt(bsr_df[utc_col]) if utc_col else pd.NaT
    )
    bsr_df["BSR_Local_Date"] = (
        normalize_dt(bsr_df[local_col]) if local_col else pd.NaT
    )

    utc_in_range = bsr_df["BSR_UTC_Date"].between(start_ts, end_ts)
    local_in_range = bsr_df["BSR_Local_Date"].between(start_ts, end_ts)

    bsr_df["Within_Period_OK"] = utc_in_range | local_in_range
    bsr_df["Within_Period_Remark"] = bsr_df["Within_Period_OK"].apply(
        lambda x: "" if x else "Date outside monitoring period"
    )

    return bsr_df

# ----------------------------- 4️⃣ Completeness Check -----------------------------
def get_sport_from_rosco(rosco_path):
    try:
        df_rosco = pd.read_excel(rosco_path, sheet_name="General Information", header=None)

        for i in range(len(df_rosco)):
            for j in range(len(df_rosco.columns) - 1):
                cell = str(df_rosco.iat[i, j]).strip().lower()

                if cell == "sports":
                    return str(df_rosco.iat[i, j + 1]).strip().lower()

        return ""
    except:
        return ""


def is_motorsport_type(sport_value):
    keywords = ["motor", "motorsport", "formula", "f1", "moto", "nascar", "race"]
    return any(k in sport_value for k in keywords)

def completeness_check(df, bsr_cols, rules, rosco_path= None):
    colmap = {
        "tv_channel": _find_column(df, bsr_cols['tv_channel']),
        "channel_id": _find_column(df, bsr_cols.get('channel_id')),
        "broadcaster": _find_column(df, bsr_cols.get('broadcaster')),
        "type_of_program": _find_column(df, bsr_cols.get('type_of_program')),
        "match_day": _find_column(df, bsr_cols.get('matchday') or bsr_cols.get('match_day', [])),
        "home_team": _find_column(df, bsr_cols.get('home_team')),
        "away_team": _find_column(df, bsr_cols.get('away_team')),
        "aud_estimates": _find_column(df, bsr_cols.get('aud_estimates')),
        "aud_metered": _find_column(df, bsr_cols.get('aud_metered')),
        "source": _find_column(df, bsr_cols.get('source'))
    }

    df["Completeness_OK"] = True
    df["Completeness_Remark"] = ""

    live_types = set(rules.get('live_types', ['live', 'repeat', 'delayed']))
    sport_value = get_sport_from_rosco(rosco_path) if rosco_path else ""
    is_motorsport = is_motorsport_type(sport_value)

    for idx, row in df.iterrows():
        missing = []

        # ---------------- Base mandatory fields ----------------
        for logical, display in [
            ("tv_channel", "TV Channel"),
            ("channel_id", "Channel ID"),
            ("broadcaster", "Broadcaster"),
            ("match_day", "Match Day"),
            ("source", "Source")
        ]:
            colname = colmap.get(logical)
            if colname is None:
                missing.append(f"{display} (column not found)")
            elif not _is_present(row.get(colname)):
                missing.append(display)

        # ---------------- Audience logic ----------------
        aud_est_col = colmap.get("aud_estimates")
        aud_met_col = colmap.get("aud_metered")

        if not aud_est_col and not aud_met_col:
            missing.append("Audience (Estimates/Metered) (columns not found)")
        else:
            est_present = _is_present(row.get(aud_est_col)) if aud_est_col else False
            met_present = _is_present(row.get(aud_met_col)) if aud_met_col else False

            if not est_present and not met_present:
                missing.append("Both Audience fields are empty")
            elif est_present and met_present:
                missing.append("Both Audience fields are filled")

        # ---------------- Program type ----------------
        type_col = colmap.get("type_of_program")
        prog_type = str(row.get(type_col) or "").strip().lower() if type_col else ""

        # ---------------- Simulcast detection ----------------
        is_simulcast = False
        for value in row.values:
            try:
                if "simulcast" in str(value).lower():
                    is_simulcast = True
                    break
            except Exception:
                continue

        # ---------------- Home / Away logic ----------------
        home_col = colmap.get("home_team")
        away_col = colmap.get("away_team")

        # ONLY enforce for Live / Repeat / Delayed AND not simulcast
        if prog_type in live_types and not is_simulcast and not is_motorsport:
            if not home_col:
                missing.append("Home Team (column not found)")
            elif not _is_present(row.get(home_col)):
                missing.append("Home Team")

            if not away_col:
                missing.append("Away Team (column not found)")
            elif not _is_present(row.get(away_col)):
                missing.append("Away Team")

        # Highlights & Magazine & Support → no Home/Away checks at all

        # ---------------- Final result ----------------
        if missing:
            df.at[idx, "Completeness_OK"] = False
            df.at[idx, "Completeness_Remark"] = "; ".join(missing)
        else:
            df.at[idx, "Completeness_Remark"] = "All key fields present"

    return df


# ----------------------------- 5️⃣ Overlap / Duplicate / Day Break -----------------------------
def overlap_duplicate_daybreak_check(df, bsr_cols, rules):
    df = df.copy()

    # --------------------------------------------------
    # Column resolution
    # --------------------------------------------------
    col_channel = _find_column(df, bsr_cols.get("tv_channel"))
    col_channel_id = _find_column(df, bsr_cols.get("channel_id"))
    col_market = _find_column(df, bsr_cols.get("market"))
    col_broadcaster = _find_column(df, bsr_cols.get("broadcaster"))

    col_date = (
        _find_column(df, ["Date (UTC)", "Date (UTC/GMT)"])
        or _find_column(df, ["Date"])
    )
    col_start = (
        _find_column(df, ["Start (UTC)", "Start UTC"])
        or _find_column(df, ["Start"])
    )
    col_end = (
        _find_column(df, ["End (UTC)", "End UTC"])
        or _find_column(df, ["End"])
    )

    col_prog_type = (
        _find_column(df, bsr_cols.get("type_of_program"))
        or _find_column(df, ["Program Type", "Type of Program"])
    )

    if not col_market or not col_date or not col_start or not col_end:
        return df

    compare_channel = col_channel if col_channel else col_channel_id
    if not compare_channel:
        return df

    # --------------------------------------------------
    # Build timezone-naive datetimes ONLY
    # --------------------------------------------------
    def build_dt(d, t):
        if pd.isna(d) or pd.isna(t):
            return pd.NaT
        try:
            date_part = pd.to_datetime(d, errors="coerce")
            if pd.isna(date_part):
                return pd.NaT

            # 🔥 HANDLE EXCEL TIME (0 / 1 issue FIX)
            if isinstance(t, (int, float)):
                time_part = pd.to_timedelta(float(t), unit="D")
            else:
                # Handle HH:MM:SS
                try:
                    time_part = pd.to_timedelta(str(t))
                except:
                    return pd.NaT

            return date_part + time_part

        except Exception:
            return pd.NaT

    df["_start_dt"] = df.apply(lambda r: build_dt(r[col_date], r[col_start]), axis=1)
    df["_end_dt"]   = df.apply(lambda r: build_dt(r[col_date], r[col_end]), axis=1)

    # HARD safety: ensure tz-naive
    df["_start_dt"] = pd.to_datetime(df["_start_dt"], errors="coerce").dt.tz_localize(None)
    df["_end_dt"]   = pd.to_datetime(df["_end_dt"], errors="coerce").dt.tz_localize(None)

    # --------------------------------------------------
    # Fix cross-midnight programs
    # --------------------------------------------------
    mask_midnight = (
        pd.notna(df["_start_dt"]) &
        pd.notna(df["_end_dt"]) &
        (df["_end_dt"] < df["_start_dt"])
    )
    df.loc[mask_midnight, "_end_dt"] += pd.Timedelta(days=1)

    # --------------------------------------------------
    # Normalize program type
    # --------------------------------------------------
    df["_prog_type_norm"] = (
        df[col_prog_type].fillna("").astype(str).str.lower().str.strip()
        if col_prog_type else ""
    )

    # --------------------------------------------------
    # Preserve original order
    # --------------------------------------------------
    df["_orig_idx"] = df.index

    # --------------------------------------------------
    # ✅ CRITICAL: FULL CHRONOLOGICAL SORT
    # Channel → Market → Date → Start time
    # --------------------------------------------------
    df["_sort_date"] = df["_start_dt"].dt.date

    df = df.sort_values(
        [compare_channel, col_market, "_sort_date", "_start_dt"],
        na_position="last"
    ).reset_index(drop=True)

    n = len(df)

    # --------------------------------------------------
    # Output containers
    # --------------------------------------------------
    overlap_ok   = [pd.NA] * n
    overlap_r    = [""] * n
    duplicate_ok = [True] * n
    duplicate_r  = [""] * n
    daybreak_ok  = [pd.NA] * n
    daybreak_r   = [""] * n

    # --------------------------------------------------
    # Duplicate check (IGNORE INTERNET / WWW)
    # --------------------------------------------------
    col_pay_free = _find_column(df, ["Pay/Free TV", "Pay Free TV", "Platform", "Distribution"])

    dup_cols = [compare_channel, col_market, col_date, col_start, col_end]
    if col_broadcaster:
        dup_cols.insert(2, col_broadcaster)

    # Default: no duplicates
    dup_mask = pd.Series([False] * n)

    try:
        if col_pay_free:
            # Identify internet / www rows (case-insensitive, partial match)
            internet_mask = (
                df[col_pay_free]
                .fillna("")
                .astype(str)
                .str.lower()
                .str.contains(r"internet|internet stream|www", regex=True)
            )

            # Run duplicate check ONLY on non-internet rows
            non_internet_df = df.loc[~internet_mask, dup_cols]
            dup_non_internet = non_internet_df.duplicated(keep=False)

            # Map results back to full DataFrame
            dup_mask.loc[~internet_mask] = dup_non_internet.values
        else:
            # No Pay/Free TV column → original behavior
            dup_mask = df.duplicated(subset=dup_cols, keep=False)

    except Exception:
        dup_mask = pd.Series([False] * n)

    # Assign results
    for i in range(n):
        if dup_mask.iloc[i]:
            duplicate_ok[i] = False
            duplicate_r[i] = "In-market duplicate (same channel/market/date/start/end)"

    # --------------------------------------------------
    # ✅ OVERLAP CHECK (WITH INTERNET / MATCH BYPASS)
    # --------------------------------------------------
    VALID_TYPES = {"live", "repeat", "delayed"}

    col_pay_free = _find_column(df, ["Pay/Free TV", "Pay Free TV", "Platform", "Distribution"])
    col_combined = _find_column(df, ["Combined"])
    col_phase = _find_column(df, ["Phase/Fixture/Episode", "Phase / Fixture / Episode"])
    col_prog_desc = _find_column(df, ["Program Description", "Program Desc"])

    def normalize_match_text(text):
        if not text:
            return ""
        text = str(text).lower()
        text = re.sub(r"(simulcast|live|repeat|delayed)", "", text)
        text = re.sub(r"\bvs\.?\b|\bv\b", "vs", text)
        text = re.sub(r"[^a-z0-9\s]", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text

    def get_match_signature(row):
        parts = []
        if col_combined:
            parts.append(str(row[col_combined] or ""))
        if col_phase:
            parts.append(str(row[col_phase] or ""))
        if col_prog_desc:
            parts.append(str(row[col_prog_desc] or ""))
        return normalize_match_text(" ".join(parts))

    def is_internet_row(row):
        if not col_pay_free:
            return False
        val = str(row[col_pay_free] or "").lower()
        return "internet" in val or "www" in val

    prev_key = None
    prev_end = None
    prev_row = None
    prev_match = None

    for i in range(n):
        row = df.iloc[i]
        key = (row[compare_channel], row[col_market], row["_sort_date"])
        start = row["_start_dt"]
        end = row["_end_dt"]
        ptype = row["_prog_type_norm"]

        # --------------------------------------------------
        # 🚀 HARD INTERNET BYPASS
        # --------------------------------------------------
        if is_internet_row(row):
            overlap_ok[i] = True
            overlap_r[i] = "Internet Channel, skipped overlap check"
            prev_key = key
            prev_end = end
            prev_row = row
            prev_match = get_match_signature(row)
            continue

        if key != prev_key:
            prev_end = None
            prev_row = None
            prev_match = None

        if ptype and ptype not in VALID_TYPES:
            overlap_ok[i] = pd.NA
            overlap_r[i] = f"Ignored program type '{ptype}'"
            prev_key = key
            continue

        if pd.isna(start) or pd.isna(end) or end <= start:
            overlap_ok[i] = pd.NA
            overlap_r[i] = "Not Applicable – invalid timing"
            prev_key = key
            continue

        if prev_end is None:
            overlap_ok[i] = True
            overlap_r[i] = "OK (first program)"
            prev_end = end
            prev_row = row
            prev_match = get_match_signature(row)
            prev_key = key
            continue

        if start == prev_end:
            overlap_ok[i] = True
            overlap_r[i] = "OK – back-to-back"
            prev_end = end
            prev_row = row
            prev_match = get_match_signature(row)
            prev_key = key
            continue

        if start < prev_end:
            overlap_ok[i] = False
            overlap_r[i] = (
                f"Overlap: starts {start.time()} "
                f"before previous ends {prev_end.time()}"
            )
            prev_end = max(prev_end, end)
            prev_row = row
            prev_match = get_match_signature(row)
            prev_key = key
            continue

        overlap_ok[i] = True
        overlap_r[i] = "OK"
        prev_end = end
        prev_row = row
        prev_match = get_match_signature(row)
        prev_key = key

    # --------------------------------------------------
    # ✅ FINAL DAYBREAK LOGIC (FIXED)
    # --------------------------------------------------
    gap_tolerance = rules.get("daybreak_gap_tolerance_min", 30)

    for i in range(1, n):
        prev = df.iloc[i - 1]
        curr = df.iloc[i]

        daybreak_ok[i] = True
        daybreak_r[i] = "OK"

        if not (
            str(prev[compare_channel]) == str(curr[compare_channel]) and
            str(prev[col_market]) == str(curr[col_market])
        ):
            continue

        if pd.isna(prev["_end_dt"]) or pd.isna(curr["_start_dt"]):
            daybreak_r[i] = "OK – missing timestamps"
            continue

        prev_end = prev["_end_dt"]
        curr_start = curr["_start_dt"]

        prev_combined = str(prev.get(col_combined, "")).strip().lower()
        curr_combined = str(curr.get(col_combined, "")).strip().lower()

        if not prev_combined or not curr_combined or prev_combined != curr_combined:
            daybreak_r[i] = "OK – different match"
            continue

        # ✅ FIXED DATE LOGIC
        same_or_next_day = (
            curr_start.date() == prev_end.date() or
            curr_start.date() == (prev_end.date() + pd.Timedelta(days=1))
        )

        if not same_or_next_day:
            daybreak_r[i] = "OK – not a daybreak"
            continue

        gap = (curr_start - prev_end).total_seconds() / 60

        if 0 <= gap <= gap_tolerance:
            daybreak_ok[i] = False
            daybreak_r[i] = "Daybreak – same match continued across midnight"
        else:
            daybreak_r[i] = f"OK – gap too large ({gap:.1f} min)"

    # --------------------------------------------------
    # Final assignment
    # --------------------------------------------------
    df["Duplicate_OK"] = duplicate_ok
    df["Duplicate_Remark"] = duplicate_r
    df["Overlap_OK"] = overlap_ok
    df["Overlap_Remark"] = overlap_r
    df["Daybreak_OK"] = daybreak_ok
    df["Daybreak_Remark"] = daybreak_r

    return df.sort_values("_orig_idx").drop(
        columns=[
            "_start_dt",
            "_end_dt",
            "_sort_date",
            "_orig_idx",
            "_prog_type_norm"
        ],
        errors="ignore"
    )

# --------------------------------------------------
# 6️⃣ Program Category Check 
# --------------------------------------------------
def program_category_check(bsr_path, df, col_map, rules, file_rules):
    from datetime import datetime, timedelta, time

    # -------------------------
    # Helpers
    # -------------------------
    def find_col(names):
        for c in df.columns:
            c_clean = c.strip().lower()
            for n in names:
                if n.strip().lower() in c_clean:
                    return c
        return None

    def parse_date(val):
        if pd.isna(val):
            return None
        if isinstance(val, (datetime, pd.Timestamp)):
            return val.date()
        return pd.to_datetime(val,errors="coerce").date()
    
    def normalize_team(name):
        if pd.isna(name):
            return ""
        name = str(name).lower()
        # Remove everything in parentheses including the parentheses
        name = re.sub(r"\(.*?\)", "", name)
        # Remove common suffixes that cause mismatches
        name = name.replace("fsf", "").replace("fs", "").replace("at.", "atletico")
        # Keep only alphanumeric
        name = re.sub(r"[^a-z0-9]", "", name)
        return name.strip()

    def parse_time(val):
        if pd.isna(val):
            return None
        if isinstance(val, time):
            return val
        if isinstance(val, (datetime, pd.Timestamp)):
            return val.time()
        if isinstance(val, str) and not val.strip().startswith("#"):
            return pd.to_datetime(val, errors="coerce").time()
        return None

    # -------------------------
    # Column detection
    # -------------------------
    col_program_type = find_col(["program type", "type of program"])
    col_desc = find_col(["combined (translated)", "program description", "description"])
    col_duration = find_col(["duration", "duration (mins)"])
    # ---- Robust UTC column detection ----
    col_datetime_utc = find_col(["date + time utc", "datetime utc", "date time utc"])
    if col_datetime_utc:
        col_date_utc = col_datetime_utc
        col_start_utc = col_datetime_utc
    else:
        col_date_utc = find_col(["date"])
        col_start_utc = find_col(["start"])
    col_home = find_col(["home team"])
    col_away = find_col(["away team"])
    col_phase = find_col(["phase", "fixture", "episode"])
    
    # -------------------------
    # Extract Monitoring Period (C3 of ROSCO)
    # -------------------------
    monitor_start = None
    monitor_end = None

    rosco_path = file_rules.get("rosco_path")

    if rosco_path:
        try:
            rosco_df = pd.read_excel(rosco_path, header=None)
            cell_value = str(rosco_df.iloc[2, 2])  # C3
            dates = re.findall(r"\d{2}[-/]\d{2}[-/]\d{4}", cell_value)

            if len(dates) == 2:
                monitor_start = pd.to_datetime(dates[0], dayfirst=True).date()
                monitor_end = pd.to_datetime(dates[1], dayfirst=True).date()
        except Exception:
            pass
    # -------------------------
    # Keywords
    # -------------------------
    highlight_re = re.compile(
        r"\b(hits|hl|highlights|hlts|overview|review|show|goals?|summary|specials|league|reload)\b",
        re.I
    )

    magazine_re = re.compile(
        r"\b(sports|show|league|magazine|support|studio|magazin|weekly|preview|analysis|review|specials|weekly new|coming soon|coming|pre|post|Chrcha|interview)\b",
        re.I
    )

    # -------------------------
    # Tolerances
    # -------------------------
    live_tol_min = rules.get("live_tolerance_min")
    live_tolerance = timedelta(minutes=int(live_tol_min)) if live_tol_min else timedelta(minutes=60)

    highlight_tol_min = rules.get("highlight_tolerance_min")
    highlight_tolerance_min = int(highlight_tol_min) if highlight_tol_min not in [None, "", 0] else None

    # -------------------------
    # Load Fixtures (Fixed for "Fixtures list")
    # -------------------------
    fixtures_df = None

    try:
        xl = pd.ExcelFile(bsr_path)

        for s in xl.sheet_names:
            if "fixtures" in s.strip().lower():
                fixtures_df = xl.parse(s)
                break

    except Exception:
        fixtures_df = None

    # -------------------------
    # Precompute BSR UTC start
    # -------------------------
    df["_bsr_start_utc"] = None

    for i, r in df.iterrows():

        # Case 1: Combined UTC datetime column exists
        if col_date_utc == col_start_utc:
            raw_val = r.get(col_date_utc)

            if pd.notna(raw_val):
                try:
                    df.at[i, "_bsr_start_utc"] = pd.to_datetime(raw_val, errors="coerce")
                except:
                    df.at[i, "_bsr_start_utc"] = None

        # Case 2: Separate date and time columns
        else:
            d = parse_date(r.get(col_date_utc))
            t = parse_time(r.get(col_start_utc))

            if d and t:
                df.at[i, "_bsr_start_utc"] = datetime.combine(d, t)

    # -------------------------
    # First broadcast map (Monitoring Period Only)
    # -------------------------
    first_broadcast = {}

    for _, r in df.iterrows():
        bsr_start = r["_bsr_start_utc"]
        if not bsr_start:
            continue

        bsr_date = bsr_start.date()

        if monitor_start and monitor_end:
            if not (monitor_start <= bsr_date <= monitor_end):
                continue

        key = (
            str(r.get(col_home)).strip().lower(),
            str(r.get(col_away)).strip().lower()
        )

        if key not in first_broadcast or bsr_start < first_broadcast[key]:
            first_broadcast[key] = bsr_start

    # -------------------------
    # Output columns
    # -------------------------
    df["program_category_check_result"] = ""
    df["program_category_check_remark"] = ""

    # -------------------------
    # Validation
    # -------------------------
    for idx, row in df.iterrows():
        ptype = str(row[col_program_type]).strip().lower()
        desc = str(row[col_desc]) if col_desc else ""
        bsr_start = row["_bsr_start_utc"]

        # ===== HIGHLIGHTS =====
        if ptype == "highlights":
            # Clean duration properly
            dur = None
            raw_dur = row.get(col_duration)

            if pd.notna(raw_dur):
                try:
                    # Extract numeric part only
                    dur = float(re.findall(r"\d+\.?\d*", str(raw_dur))[0])
                except Exception:
                    dur = None

            # Case 1: User has provided highlight tolerance → enforce duration
            if highlight_tolerance_min is not None:
                if dur is None:
                    df.at[idx, "program_category_check_result"] = "False"
                    df.at[idx, "program_category_check_remark"] = "Highlight duration missing"

                elif dur <= highlight_tolerance_min:
                    df.at[idx, "program_category_check_result"] = "True"
                    df.at[idx, "program_category_check_remark"] = (
                        f"Valid Highlight (duration ≤ {highlight_tolerance_min} mins)"
                    )

                else:
                    df.at[idx, "program_category_check_result"] = "False"
                    df.at[idx, "program_category_check_remark"] = (
                        f"Highlight duration exceeds {highlight_tolerance_min} mins"
                    )

            # Case 2: User did NOT provide tolerance → bypass duration check
            else:
                df.at[idx, "program_category_check_result"] = "True"
                df.at[idx, "program_category_check_remark"] = (
                    "Valid Highlights program (duration check not applied)"
                )

        # ===== MAGAZINE & SUPPORT =====
        elif ptype in ["magazine & support", "magazine and support"]:

            # Always True irrespective of keywords
            df.at[idx, "program_category_check_result"] = "True"

            if magazine_re.search(desc):
                df.at[idx, "program_category_check_remark"] = "Valid Magazine & Support (keywords present)"
            else:
                df.at[idx, "program_category_check_remark"] = "Valid Magazine & Support"

        # ===== LIVE =====
        elif ptype == "live":

            # --- STRICT SIMULCAST OVERRIDE (PHASE COLUMN ONLY) ---
            if col_phase is not None:
                phase_val = str(row[col_phase]).strip().lower()
                if "simulcast" in phase_val:
                    df.at[idx, "program_category_check_result"] = "True"
                    df.at[idx, "program_category_check_remark"] = "Valid Live program (Simulcast)"
                    continue

            # Check if bsr_start exists
            bsr_start = row.get("_bsr_start_utc")
            if fixtures_df is None or pd.isna(bsr_start):
                df.at[idx, "program_category_check_result"] = "False"
                df.at[idx, "program_category_check_remark"] = "Invalid Live timing or fixtures missing"
                continue

            # Ensure bsr_start is naive (no timezone) for comparison
            if hasattr(bsr_start, 'tzinfo') and bsr_start.tzinfo is not None:
                bsr_start = bsr_start.replace(tzinfo=None)

            home = str(row[col_home]).strip().lower() if col_home else ""
            away = str(row[col_away]).strip().lower() if col_away else ""

            matched = False
            for _, fx in fixtures_df.iterrows():
                try:
                    # 1. Parse Fixture Start
                    fx_raw_start = fx.get("Date + Time UTC")
                    fx_start = pd.to_datetime(fx_raw_start, errors="coerce")
                    if pd.isna(fx_start):
                        continue
                    
                    if hasattr(fx_start, 'tzinfo') and fx_start.tzinfo is not None:
                        fx_start = fx_start.replace(tzinfo=None)

                    # 2. Parse Duration and Calculate End
                    raw_fx_dur = fx.get("Duration")
                    try:
                        if isinstance(raw_fx_dur, (time, datetime)):
                            dur_delta = timedelta(hours=raw_fx_dur.hour, minutes=raw_fx_dur.minute, seconds=raw_fx_dur.second)
                        else:
                            # Handles "02:00:00" or similar strings
                            dur_delta = pd.to_timedelta(str(raw_fx_dur))
                        
                        fx_end = fx_start + dur_delta
                    except:
                        fx_end = fx_start + timedelta(hours=3) # Fallback to 3h for Live

                    # 3. Team Matching (Normalized)
                    h_match = normalize_team(home) == normalize_team(fx.get("Home Team", ""))
                    a_match = normalize_team(away) == normalize_team(fx.get("Away Team", ""))

                    # 4. Time Window Check
                    # We check if BSR start falls within Fixture Start - Tolerance AND Fixture End + Tolerance
                    time_match = (fx_start - live_tolerance <= bsr_start <= fx_end + live_tolerance)

                    if h_match and a_match and time_match:
                        matched = True
                        break
                except Exception as e:
                    # Log the specific error if needed: print(f"Error at index {idx}: {e}")
                    continue

            if matched:
                df.at[idx, "program_category_check_result"] = "True"
                df.at[idx, "program_category_check_remark"] = "Valid Live program"
            else:
                df.at[idx, "program_category_check_result"] = "False"
                df.at[idx, "program_category_check_remark"] = "Live program outside tolerance or team mismatch"
        
        # =====================================================
        # DELAYED
        # =====================================================
        elif "delayed" in ptype:

            df.at[idx, "program_category_check_result"] = "True"
            df.at[idx, "program_category_check_remark"] = "Valid Delayed"

        # =====================================================
        # REPEAT
        # =====================================================
        elif "repeat" in ptype:

            df.at[idx, "program_category_check_result"] = "True"
            df.at[idx, "program_category_check_remark"] = "Valid Repeat"

        # =====================================================
        # OTHERS
        # =====================================================
        else:
            df.at[idx, "program_category_check_result"] = "NA"
            df.at[idx, "program_category_check_remark"] = "Program type not applicable"

        # # =====================================================
        # # DELAYED
        # # =====================================================
        # elif ptype == "delayed":

        #     if not bsr_start:
        #         df.at[idx, "program_category_check_result"] = "False"
        #         df.at[idx, "program_category_check_remark"] = "Invalid start time"
        #         continue

        #     home = str(row[col_home]).strip().lower()
        #     away = str(row[col_away]).strip().lower()

        #     first_time = first_broadcast.get((home, away))

        #     if first_time and bsr_start == first_time:
        #         df.at[idx, "program_category_check_result"] = "True"
        #         df.at[idx, "program_category_check_remark"] = "Valid Delayed (first in monitoring period)"
        #     else:
        #         df.at[idx, "program_category_check_result"] = "False"
        #         df.at[idx, "program_category_check_remark"] = "Not first in monitoring period - should be Repeat"

        # # =====================================================
        # # REPEAT
        # # =====================================================
        # elif ptype == "repeat":

        #     if not bsr_start:
        #         df.at[idx, "program_category_check_result"] = "False"
        #         df.at[idx, "program_category_check_remark"] = "Invalid start time"
        #         continue

        #     home = str(row[col_home]).strip().lower()
        #     away = str(row[col_away]).strip().lower()

        #     first_time = first_broadcast.get((home, away))

        #     if first_time and bsr_start > first_time:
        #         df.at[idx, "program_category_check_result"] = "True"
        #         df.at[idx, "program_category_check_remark"] = "Valid Repeat"
        #     else:
        #         df.at[idx, "program_category_check_result"] = "False"
        #         df.at[idx, "program_category_check_remark"] = "First broadcast cannot be Repeat"
        # else:
        #     df.at[idx, "program_category_check_result"] = "NA"
        #     df.at[idx, "program_category_check_remark"] = "Program type not applicable"

    df.drop(columns=["_bsr_start_utc"], inplace=True, errors="ignore")
    return df
        
# -----------------------------------------------------------
# 8️⃣ Event / Matchday / Competition Check
# -----------------------------------------------------------
import pandas as pd
import logging


def _norm(val):
    """
    Strong normalization for matching
    """

    if pd.isna(val):
        return ""

    val = str(val).strip().lower()

    # remove duplicate spaces
    val = " ".join(val.split())

    # remove trailing .0
    if val.endswith(".0"):
        val = val[:-2]

    return val


def _norm_date(val):
    """
    Normalize dates to YYYY-MM-DD
    """

    if pd.isna(val):
        return ""

    try:
        return pd.to_datetime(
            val,
            dayfirst=True
        ).strftime("%Y-%m-%d")

    except Exception:
        return _norm(val)


def check_event_matchday_competition(
    df,
    bsr_path,
    col_map,
    file_rules
):

    logging.info(
        "Starting Event / Matchday / Fixture consistency check..."
    )

    bsr_cols = col_map.get("bsr", {})
    fix_cols = col_map.get("fixture", {})

    # ---------------------------------------------------------
    # TYPE OF PROGRAM COLUMN
    # ---------------------------------------------------------
    col_progtype = _find_column(
        df,
        bsr_cols.get("type_of_program")
    )

    if not col_progtype:

        logging.error(
            "❌ Type of program column not found"
        )

        df["Event_Matchday_OK"] = False
        df["Event_Matchday_Remark"] = (
            "Type of program column missing"
        )

        return df

    # ---------------------------------------------------------
    # DEFAULT OUTPUT COLUMNS
    # ---------------------------------------------------------
    df["Event_Matchday_OK"] = pd.NA
    df["Event_Matchday_Remark"] = "Not applicable"

    CHECKABLE_TYPES = {
        "live",
        "delayed",
        "live delayed",
        "delay",
    }

    SKIP_TYPES = {
        "highlights",
        "magazine",
        "support",
    }

    # ---------------------------------------------------------
    # LOAD FIXTURE SHEET
    # ---------------------------------------------------------
    fixture_df = None

    try:

        excel_file = pd.ExcelFile(bsr_path)

        fixture_keyword = file_rules.get(
            "fixture_sheet_keyword",
            "fixture"
        )

        fixture_sheet = next(
            (
                s for s in excel_file.sheet_names
                if fixture_keyword.lower() in s.lower()
            ),
            None
        )

        if fixture_sheet:

            fixture_df = excel_file.parse(
                fixture_sheet
            )

            fixture_df.columns = [
                str(c).strip()
                for c in fixture_df.columns
            ]

            logging.info(
                f"✅ Loaded fixture sheet: {fixture_sheet}"
            )

        else:

            logging.warning(
                "⚠️ Fixture sheet not found"
            )

    except Exception as e:

        logging.error(
            f"❌ Error loading fixture sheet: {e}"
        )

    # ---------------------------------------------------------
    # FIXTURE SHEET VALIDATION
    # ---------------------------------------------------------
    if fixture_df is None:

        df["Event_Matchday_OK"] = False

        df["Event_Matchday_Remark"] = (
            "Fixture sheet missing"
        )

        return df

    # ---------------------------------------------------------
    # SAFE EPISODE MAPPING
    # ---------------------------------------------------------
    fixture_episode_mapping = (
        fix_cols.get("phase_fixture_episode")
        or fix_cols.get("episode")
        or "Phase / Fixture / Episode Desc."
    )

    bsr_episode_mapping = (
        bsr_cols.get("phase_fixture_episode")
        or bsr_cols.get("episode")
        or "Phase / Fixture / Episode Desc."
    )

    # ---------------------------------------------------------
    # RESOLVE FIXTURE COLUMNS
    # ---------------------------------------------------------
    fix_comp_col = _find_column(
        fixture_df,
        fix_cols.get("competition")
    )

    fix_matchday_col = _find_column(
        fixture_df,
        fix_cols.get("match_day")
    )

    fix_episode_col = _find_column(
        fixture_df,
        fixture_episode_mapping
    )

    # ---------------------------------------------------------
    # RESOLVE BSR COLUMNS
    # ---------------------------------------------------------
    bsr_comp_col = _find_column(
        df,
        bsr_cols.get("competition")
    )

    bsr_event_col = _find_column(
        df,
        bsr_cols.get("event")
    )

    bsr_matchday_col = _find_column(
        df,
        bsr_cols.get("match_day")
    )

    bsr_episode_col = _find_column(
        df,
        bsr_episode_mapping
    )

    logging.info(
        f"""
        Fixture cols:
        competition = {fix_comp_col}
        matchday    = {fix_matchday_col}
        episode     = {fix_episode_col}

        BSR cols:
        competition = {bsr_comp_col}
        event       = {bsr_event_col}
        matchday    = {bsr_matchday_col}
        episode     = {bsr_episode_col}
        """
    )

    # ---------------------------------------------------------
    # REQUIRED COLUMN VALIDATION
    # ---------------------------------------------------------
    required_cols = {
        "fix_comp_col": fix_comp_col,
        "fix_matchday_col": fix_matchday_col,
        "fix_episode_col": fix_episode_col,
        "bsr_matchday_col": bsr_matchday_col,
        "bsr_episode_col": bsr_episode_col,
    }

    missing = [
        k for k, v in required_cols.items()
        if not v
    ]

    if missing:

        logging.error(
            f"❌ Missing required columns: {missing}"
        )

        df["Event_Matchday_OK"] = False

        df["Event_Matchday_Remark"] = (
            f"Missing columns: {missing}"
        )

        return df

    # ---------------------------------------------------------
    # NORMALIZE FIXTURE DATA
    # ---------------------------------------------------------
    fixture_df["_competition"] = fixture_df[
        fix_comp_col
    ].apply(_norm)

    fixture_df["_matchday"] = fixture_df[
        fix_matchday_col
    ].apply(_norm_date)

    fixture_df["_episode"] = fixture_df[
        fix_episode_col
    ].apply(_norm)

    # ---------------------------------------------------------
    # CREATE FIXTURE KEY
    # ---------------------------------------------------------
    fixture_df["_fixture_key"] = (
        fixture_df["_competition"]
        + "||"
        + fixture_df["_matchday"]
        + "||"
        + fixture_df["_episode"]
    )

    fixture_keys = set(
        fixture_df["_fixture_key"]
    )

    logging.info(
        f"✅ Fixture keys created: {len(fixture_keys)}"
    )

    # ---------------------------------------------------------
    # MAIN CHECK
    # ---------------------------------------------------------
    for i, row in df.iterrows():

        try:

            prog_type = _norm(
                row.get(col_progtype, "")
            )

            # -------------------------------------------------
            # SKIP PROGRAM TYPES
            # -------------------------------------------------
            if prog_type in SKIP_TYPES:

                df.at[i, "Event_Matchday_OK"] = pd.NA

                df.at[i, "Event_Matchday_Remark"] = (
                    "Not applicable"
                )

                continue

            # -------------------------------------------------
            # NON CHECKABLE TYPES
            # -------------------------------------------------
            if prog_type not in CHECKABLE_TYPES:

                df.at[i, "Event_Matchday_OK"] = pd.NA

                df.at[i, "Event_Matchday_Remark"] = (
                    f"Not applicable ({prog_type})"
                )

                continue

            # -------------------------------------------------
            # COMPETITION / EVENT FALLBACK
            # -------------------------------------------------
            competition_value = ""

            # Prefer Competition
            if bsr_comp_col:

                competition_value = _norm(
                    row.get(bsr_comp_col, "")
                )

            # Fallback to Event
            if (
                not competition_value
                and bsr_event_col
            ):

                competition_value = _norm(
                    row.get(bsr_event_col, "")
                )

            comp = competition_value

            # -------------------------------------------------
            # MATCHDAY
            # -------------------------------------------------
            matchday = _norm_date(
                row.get(bsr_matchday_col, "")
            )

            # -------------------------------------------------
            # EPISODE
            # -------------------------------------------------
            episode = _norm(
                row.get(bsr_episode_col, "")
            )

            # -------------------------------------------------
            # EMPTY CHECK
            # -------------------------------------------------
            if not comp or not matchday or not episode:

                df.at[i, "Event_Matchday_OK"] = False

                df.at[i, "Event_Matchday_Remark"] = (
                    f"Missing values | "
                    f"competition/event='{comp}' | "
                    f"matchday='{matchday}' | "
                    f"episode='{episode}'"
                )

                continue

            # -------------------------------------------------
            # CREATE BSR KEY
            # -------------------------------------------------
            bsr_key = (
                comp
                + "||"
                + matchday
                + "||"
                + episode
            )

            # -------------------------------------------------
            # MATCH AGAINST FIXTURE
            # -------------------------------------------------
            if bsr_key in fixture_keys:

                df.at[i, "Event_Matchday_OK"] = True

                df.at[i, "Event_Matchday_Remark"] = (
                    "Fixture match"
                )

            else:

                df.at[i, "Event_Matchday_OK"] = False

                df.at[i, "Event_Matchday_Remark"] = (
                    f"No fixture match | {bsr_key}"
                )

        except Exception as e:

            logging.exception(e)

            df.at[i, "Event_Matchday_OK"] = False

            df.at[i, "Event_Matchday_Remark"] = (
                f"Error: {e}"
            )

    logging.info(
        "✅ Event / Matchday / Fixture check completed"
    )

    return df

# ── Shared helpers (define once outside the function) ───────────────────────

def _norm(val) -> str:
    """Lowercase, strip, collapse whitespace, strip time from date strings."""
    s = str(val).strip().lower()
    if s in ('nan', 'none', 'nat', ''):
        return ''
    # Strip time portion from 'yyyy-mm-dd hh:mm...' so dates compare cleanly
    s = re.sub(r'^(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}.*$', r'\1', s)
    # Collapse internal whitespace
    s = re.sub(r'\s+', ' ', s)
    return s

def _normalise_series(series) -> 'pd.Series':
    """Apply _norm to an entire Series."""
    return series.astype(str).apply(_norm)

# -----------------------------------------------------------
# 9️⃣ Market / Channel / Program / Duration Consistency Check
# -----------------------------------------------------------

def market_channel_consistency_check(df_bsr, rosco_path, col_map, file_rules):
    logging.info("🔍 Starting Market & Channel Consistency Check...")
    bsr_cols = col_map['bsr']
    rosco_cols = col_map.get('rosco', {})
    def normalize_channel(name):
        if pd.isna(name) or name is None:
            return ""
        s = str(name)
        s = re.sub(r"\(.*?\)|\[.*?\]", "", s)
        s = re.split(r"[-–—]", s)[0]
        s = re.sub(r"[^0-9a-zA-Z\s]", " ", s)
        return re.sub(r"\s+", " ", s).strip().lower()
    rosco_df = None
    if rosco_path:
        try:
            xls = pd.ExcelFile(rosco_path)
            ignore_sheet = file_rules.get('rosco_ignore_sheet', 'general')
            sheet_name = next((s for s in xls.sheet_names if ignore_sheet not in s.lower()), None)
            if sheet_name:
                rosco_df = xls.parse(sheet_name)
            else:
                logging.warning(f"No valid sheet found in ROSCO (ignoring '{ignore_sheet}').")
        except Exception as e:
            logging.error(f"Error loading ROSCO file: {e}")
            df_bsr["Market_Channel_Consistency_OK"] = False
            df_bsr["Market_Channel_Program_Remark"] = f"Error loading ROSCO: {e}"
            return df_bsr
    valid_pairs = set()
    rosco_country_col = rosco_cols.get('channel_country', 'ChannelCountry')
    rosco_name_col = rosco_cols.get('channel_name', 'ChannelName')
    if rosco_df is not None and not rosco_df.empty and {rosco_country_col, rosco_name_col}.issubset(rosco_df.columns):
        for _, row in rosco_df.iterrows():
            market = str(row[rosco_country_col]).strip().lower()
            channel = normalize_channel(row[rosco_name_col])
            if market and channel:
                valid_pairs.add((market, channel))
        logging.info(f"Loaded {len(valid_pairs)} valid Market+Channel pairs from ROSCO.")
    df_bsr["Market_Channel_Consistency_OK"] = True
    df_bsr["Market_Channel_Program_Remark"] = "OK"
    bsr_market_col = _find_column(df_bsr, bsr_cols.get('market'))
    bsr_channel_col = _find_column(df_bsr, bsr_cols.get('tv_channel'))
    if not bsr_market_col or not bsr_channel_col:
        logging.error("Market/Channel Check: BSR columns not found. Skipping.")
        df_bsr["Market_Channel_Consistency_OK"] = False
        df_bsr["Market_Channel_Program_Remark"] = "BSR columns not found"
        return df_bsr
    for idx, row in df_bsr.iterrows():
        remarks = []
        market = str(row.get(bsr_market_col, "")).strip().lower()
        channel = str(row.get(bsr_channel_col, "")).strip()
        if not market or not channel:
            df_bsr.at[idx, "Market_Channel_Consistency_OK"] = False
            remarks.append("Missing market or channel")
        elif valid_pairs:
            if (market, normalize_channel(channel)) not in valid_pairs:
                df_bsr.at[idx, "Market_Channel_Consistency_OK"] = False
                remarks.append("Market+Channel not found in ROSCO")
        df_bsr.at[idx, "Market_Channel_Program_Remark"] = "; ".join(remarks) if remarks else "OK"
    logging.info("✅ Market & Channel Consistency Check completed.")
    return df_bsr

# -----------------------------------------------------------
# 10️⃣ Domestic Market Coverage Check
# -----------------------------------------------------------
def domestic_market_check(df_worksheet, bsr_cols, monitoring_start_date=None, debug=False):
    df = df_worksheet.copy()
    df["Domestic_Market_Coverage_OK"] = True
    df["Domestic_Market_Remark"] = ""
    col_comp = _find_column(df, bsr_cols.get('competition', ['Competition']))
    col_mkt = _find_column(df, bsr_cols.get('market', ['Market']))
    col_date = _find_column(df, bsr_cols.get('date', ['Date']))
    col_prog_type = _find_column(df, bsr_cols.get('type_of_program', ['Type of Program']))
    if not all([col_comp, col_mkt, col_date, col_prog_type]):
        df["Domestic_Market_Coverage_OK"] = False
        df["Domestic_Market_Remark"] = "Skipped: Missing core BSR columns in file/config."
        return df
    DOMESTIC_MAP = {
        "premier league": ["united kingdom", "england"],
        "epl": ["united kingdom", "england"],
        "la liga": ["spain"],
        "bundesliga": ["germany", "deutschland"],
        "serie a": ["italy"],
        "ligue 1": ["france"]
    }
    monitoring_start = None
    if monitoring_start_date is not None:
        try:
            monitoring_start = pd.to_datetime(monitoring_start_date).date()
        except Exception:
            monitoring_start = None
    for idx, row in df.iterrows():
        comp = str(row.get(col_comp, "")).strip().lower()
        market = str(row.get(col_mkt, "")).strip().lower()
        date_raw = row.get(col_date)
        try:
            row_date = pd.to_datetime(date_raw).date()
        except Exception:
            row_date = None
        if monitoring_start and row_date and row_date < monitoring_start:
            continue
        domestic_markets = []
        for comp_kw, markets in DOMESTIC_MAP.items():
            if comp_kw in comp:
                domestic_markets = markets
                break
        if not domestic_markets:
            continue
        market_ok = any(dm in market for dm in domestic_markets)
        if not market_ok:
            df.at[idx, "Domestic_Market_Coverage_OK"] = False
            df.at[idx, "Domestic_Market_Remark"] = f"Missing domestic coverage. Expected one of: {domestic_markets}"
        else:
            df.at[idx, "Domestic_Market_Remark"] = "OK"
    return df

# -----------------------------------------------------------
# 11️⃣ Rates & Ratings Check
# --------------------------------------------
def rates_and_ratings_check(df, bsr_cols):
    est_col = _find_column(df, bsr_cols.get('aud_estimates'))
    met_col = _find_column(df, bsr_cols.get('aud_metered'))
    est_col_exists = est_col is not None and est_col in df.columns
    met_col_exists = met_col is not None and met_col in df.columns
    if est_col is None:
        est_col = "Audience_Estimates_Dummy"
        df[est_col] = np.nan
        logging.warning("Rates/Ratings Check: Audience Estimates column not found.")
    if met_col is None:
        met_col = "Audience_Metered_Dummy"
        df[met_col] = np.nan
        logging.warning("Rates/Ratings Check: Audience Metered column not found.")
    present_est = df[est_col].apply(_is_present)
    present_met = df[met_col].apply(_is_present)
    both_empty_mask = (~present_est) & (~present_met)
    both_present_mask = (present_est) & (present_met)
    exactly_one_mask = (present_est ^ present_met)
    df["Rates_Ratings_QC_OK"] = True
    df["Rates_Ratings_QC_Remark"] = ""
    df.loc[both_empty_mask, "Rates_Ratings_QC_OK"] = False
    df.loc[both_empty_mask, "Rates_Ratings_QC_Remark"] = "Missing audience ratings (both empty)"
    df.loc[both_present_mask, "Rates_Ratings_QC_OK"] = False
    df.loc[both_present_mask, "Rates_Ratings_QC_Remark"] = "Invalid: both metered and estimated present"
    df.loc[exactly_one_mask, "Rates_Ratings_QC_OK"] = True
    df.loc[exactly_one_mask, "Rates_Ratings_QC_Remark"] = "Valid: one rating source available"
    if est_col == "Audience_Estimates_Dummy" and est_col in df.columns:
        df.drop(columns=[est_col], inplace=True)
    if met_col == "Audience_Metered_Dummy" and met_col in df.columns:
        df.drop(columns=[met_col], inplace=True)
    return df

# -----------------------------------------------------------
# 12️⃣ Comparison of Duplicated Markets
# -----------------------------------------------------------
def duplicated_market_check(df_bsr, macro_path, project, col_map, file_rules, debug=False):

    result_col = "Duplicated_Markets_Check_OK"
    remark_col = "Duplicated_Markets_Remark"

    df_bsr[result_col] = pd.NA
    df_bsr[remark_col] = "Not Applicable"

    league_keyword = str(project.get("league_keyword", "F24 Spain")).lower()
    bsr_cols = col_map["bsr"]
    macro_cols = col_map["macro"]

    if not macro_path or not os.path.exists(macro_path):
        df_bsr[result_col] = False
        df_bsr[remark_col] = "Macro file missing"
        return df_bsr


    # -------------------------------------------------------
    #  STEP 1 — Load Excel WITHOUT trusting header_row
    # -------------------------------------------------------
    try:
        xl = pd.ExcelFile(macro_path, engine="openpyxl")

        # Pick correct sheet
        preferred = file_rules.get("macro_sheet_name", "Data Core").lower()
        sheet = next((s for s in xl.sheet_names if s.lower() == preferred), xl.sheet_names[0])

        # Read top 20 rows without header
        tmp = pd.read_excel(macro_path, sheet_name=sheet, header=None, nrows=20, dtype=str)

        required_cols = ["Projects", "Orig Market", "Orig Channel", "Dup Market", "Dup Channel"]

        header_row_index = None

        #  Find the row where all required column names appear
        for i in range(len(tmp)):
            row_vals = [str(x).strip().lower() for x in list(tmp.iloc[i].values)]
            if all(any(req.lower() == val for val in row_vals) for req in required_cols):
                header_row_index = i
                break

        if header_row_index is None:
            df_bsr[result_col] = False
            df_bsr[remark_col] = "Could not locate header row in macro file."
            return df_bsr

        # Now correctly load macro_df using detected header row
        macro_df = pd.read_excel(
            macro_path,
            sheet_name=sheet,
            header=header_row_index,
            dtype=str,
            engine="openpyxl"
        )

        macro_df.columns = [str(c).strip() for c in macro_df.columns]

    except Exception as e:
        df_bsr[result_col] = False
        df_bsr[remark_col] = f"Macro load error: {e}"
        return df_bsr


    # -------------------------------------------------------
    #  STEP 2 — Find required columns reliably
    # -------------------------------------------------------
    def find_col(df, key):
        if isinstance(key, list):
            candidates = key
        else:
            candidates = [key]

        lower = {c.lower(): c for c in df.columns}
        for cand in candidates:
            c = str(cand).strip().lower()
            if c in lower:
                return lower[c]
        return None

    proj_col = find_col(macro_df, macro_cols["projects"])
    orig_mkt_col = find_col(macro_df, macro_cols["orig_market"])
    orig_ch_col = find_col(macro_df, macro_cols["orig_channel"])
    dup_mkt_col = find_col(macro_df, macro_cols["dup_market"])
    dup_ch_col = find_col(macro_df, macro_cols["dup_channel"])

    missing = [col for col in [proj_col, orig_mkt_col, orig_ch_col, dup_mkt_col, dup_ch_col] if col is None]
    if missing:
        df_bsr[result_col] = False
        df_bsr[remark_col] = "Macro file columns not found (after auto-detect)."
        return df_bsr


    # -------------------------------------------------------
    #  STEP 3 — Filter by project keyword
    # -------------------------------------------------------
    macro_df = macro_df[
        macro_df[proj_col].astype(str).str.lower().str.contains(league_keyword, na=False)
    ]

    if macro_df.empty:
        df_bsr[result_col] = pd.NA
        df_bsr[remark_col] = f"No duplication rules found for {league_keyword}"
        return df_bsr


    # -------------------------------------------------------
    #  STEP 4 — Run duplication checks (unchanged logic)
    # -------------------------------------------------------
    mkt_col = find_col(df_bsr, bsr_cols["market"])
    ch_col = find_col(df_bsr, bsr_cols["tv_channel"])
    comp_col = find_col(df_bsr, bsr_cols["competition"])
    evt_col = find_col(df_bsr, bsr_cols["event"])

    in_league = (
        df_bsr[comp_col].astype(str).str.lower().str.contains(league_keyword, na=False)
        | df_bsr[evt_col].astype(str).str.lower().str.contains(league_keyword, na=False)
    )

    df_bsr.loc[~in_league, result_col] = pd.NA
    df_bsr.loc[~in_league, remark_col] = "Not Applicable"

    df_league = df_bsr[in_league].copy()

    for _, r in macro_df.iterrows():
        orig_market = str(r[orig_mkt_col]).strip().lower()
        orig_channel = str(r[orig_ch_col]).strip().lower()
        dup_market = str(r[dup_mkt_col]).strip().lower()
        dup_channel = str(r[dup_ch_col]).strip().lower()

        orig_rows = df_league[
            (df_league[mkt_col].str.lower() == orig_market) &
            (df_league[ch_col].str.lower() == orig_channel)
        ]
        dup_rows = df_league[
            (df_league[mkt_col].str.lower() == dup_market) &
            (df_league[ch_col].str.lower() == dup_channel)
        ]

        orig_events = set(orig_rows[evt_col].dropna().str.lower().str.strip())
        dup_events = set(dup_rows[evt_col].dropna().str.lower().str.strip())

        if not orig_events:
            status = pd.NA
            remark = f"No events found for {orig_market}/{orig_channel}"
        elif orig_events.issubset(dup_events):
            status = True
            remark = f"All {len(orig_events)} events duplicated"
        else:
            missing = orig_events - dup_events
            status = False
            remark = f"Missing {len(missing)} events"

        mask = (
            (df_bsr[mkt_col].str.lower() == orig_market) &
            (df_bsr[ch_col].str.lower() == orig_channel) &
            in_league
        ) | (
            (df_bsr[mkt_col].str.lower() == dup_market) &
            (df_bsr[ch_col].str.lower() == dup_channel) &
            in_league
        )

        df_bsr.loc[mask, result_col] = status
        df_bsr.loc[mask, remark_col] = remark

    return df_bsr
# -----------------------------------------------------------
# 13️⃣ Country & Channel IDs Check
# -----------------------------------------------------------
def country_channel_id_check(df, bsr_cols):
    """
    Check consistency of channel IDs per (market, tv_channel) pair.

    RULE:
    - For each (Market, TV-Channel) pair → must have exactly ONE unique non-blank Channel ID.
    - If same pair appears with different non-blank Channel IDs → inconsistent.
    - If the only channel_id is blank → inconsistent (Missing channel ID).
    - Same TV-Channel across different markets is allowed (treated independently).

    Adds these columns:
        Market_Channel_ID_OK (bool)
        Market_Channel_ID_Remark (str)
    """

    df = df.copy()  # work on a copy to avoid side-effects
    df["Market_Channel_ID_OK"] = True
    df["Market_Channel_ID_Remark"] = "OK"

    # Resolve column names (use _find_column)
    ch_col = _find_column(df, bsr_cols.get("tv_channel"))
    ch_id_col = _find_column(df, bsr_cols.get("channel_id"))
    mkt_col = _find_column(df, bsr_cols.get("market"))

    if not all([ch_col, ch_id_col, mkt_col]):
        logging.warning("Country/Channel ID Check: Missing required columns. Skipping.")
        df["Market_Channel_ID_OK"] = False
        df["Market_Channel_ID_Remark"] = "Check skipped: missing required columns"
        return df

    def norm(x):
        if pd.isna(x) or x is None:
            return ""
        return str(x).strip()

    # Build mapping: (market, tv_channel) → set(channel_ids) & row indices
    pair_ids = {}
    pair_idxs = {}

    for idx, row in df.iterrows():
        market = norm(row.get(mkt_col, ""))
        channel = norm(row.get(ch_col, ""))
        channel_id = norm(row.get(ch_id_col, ""))

        # Normalize for comparisons (lower-case for market and channel)
        key = (market.lower(), channel.lower())

        pair_ids.setdefault(key, set()).add(channel_id)
        pair_idxs.setdefault(key, []).append(idx)

    # Evaluate each pair
    for key, id_set in pair_ids.items():
        idxs = pair_idxs.get(key, [])
        # consider only non-blank IDs for uniqueness check
        non_blank_ids = {cid for cid in id_set if cid != ""}

        inconsistent = False
        remark = "OK"

        if len(non_blank_ids) == 0:
            inconsistent = True
            remark = "Missing channel ID"
        elif len(non_blank_ids) > 1:
            inconsistent = True
            # keep blanks visible as <BLANK> if present
            ids_list = [cid if cid != "" else "<BLANK>" for cid in sorted(id_set)]
            # include market/channel for clarity in remark
            market_display, channel_display = key
            remark = f"Conflicting Channel IDs for {channel_display} in market {market_display}: {', '.join(ids_list)}"
        else:
            inconsistent = False
            remark = "OK"

        for i in idxs:
            df.at[i, "Market_Channel_ID_OK"] = not inconsistent
            df.at[i, "Market_Channel_ID_Remark"] = remark

    return df

# -----------------------------------------------------------
# 14️⃣ Home vs Away vs Phase Consistency Check (Updated Logic)
# -----------------------------------------------------------
def home_away_vs_phase_check(df, col_map):

    result_col = "Home_vs_Away_vs_Phase_OK"
    df[result_col] = "NA"

    # Extract mapping safely
    if isinstance(col_map.get("bsr"), dict):
        b = col_map["bsr"]
    else:
        b = col_map

    def _get_best_col(df, config_key, fallbacks):
        search_terms = []

        config_val = b.get(config_key, [])
        if isinstance(config_val, list):
            search_terms.extend(config_val)
        elif isinstance(config_val, str):
            search_terms.append(config_val)

        search_terms.extend(fallbacks)

        for term in search_terms:
            if not term:
                continue
            if term in df.columns:
                return term
            for actual_col in df.columns:
                if str(actual_col).strip().lower() == str(term).strip().lower():
                    return actual_col
        return None

    # Detect Columns
    col_home = _get_best_col(df, "home_team",
                             ["Home Team", "Home", "Team 1", "Team A", "Home_Team"])

    col_away = _get_best_col(df, "away_team",
                             ["Away Team", "Away", "Team 2", "Team B", "Away_Team"])

    col_phase = _get_best_col(df, "phase_fixture_episode",
                              ["PhaseFixtureEpisode", "Phase", "Fixture",
                               "Combined (translated)", "Event", "Description", "Program"])

    col_indices = {c: i for i, c in enumerate(df.columns)}

    def clean_text(text):
        if pd.isna(text):
            return ""
        return re.sub(r"\s+", " ", str(text).strip().lower())
    
    def normalize_team(team):
        team = re.sub(r"\(.*?\)","", str(team))
        return clean_text(team)

    # Process Rows
    for idx, row in df.iterrows():

        # Column validation
        if not col_home or not col_phase:
            df.at[idx, result_col] = "Error: Required columns missing"
            continue

        home_team = clean_text(row.get(col_home, ""))
        phase_val = clean_text(row.get(col_phase, ""))

        # If no home team → Not Applicable
        if not home_team:
            df.at[idx, result_col] = "Not Applicable"
            continue

        # Detect Away Team
        away_team = ""

        if col_away:
            away_team = clean_text(row.get(col_away, ""))

        # Fallback positional logic (Home | Vs | Away)
        if not away_team:
            home_idx = col_indices.get(col_home)
            if home_idx is not None and home_idx + 2 < len(df.columns):
                sep_val = clean_text(row.iloc[home_idx + 1])
                if sep_val in ["vs", "v", "vs.", "-", "x", "v."]:
                    away_team = clean_text(row.iloc[home_idx + 2])

        # If either team missing → Not Applicable
        if not home_team or not away_team:
            df.at[idx, result_col] = "Not Applicable"
            continue
        home_team_n = normalize_team(home_team)
        away_team_n = normalize_team(away_team)
        phase_val_n = normalize_team(phase_val)

        # Final Validation
        if home_team_n in phase_val_n and away_team_n in phase_val_n:
            df.at[idx, result_col] = "OK"
        else:
            df.at[idx, result_col] = "Home/Away teams do not match PhaseFixtureEpisode"

    return df

# -----------------------------------------------------------
# 15️⃣ Multiple Live Match Consistency Check (Corrected)
# -----------------------------------------------------------

def multiple_live_match_check(df, col_map):

    result_col = "Multiple_Live_Match_OK"

    df = df.copy()

    # Default value
    df[result_col] = "NA"

    # -------------------------------------------------------
    # Safe config mapping
    # -------------------------------------------------------
    b = col_map.get("bsr", {}) if isinstance(col_map.get("bsr"), dict) else col_map

    # -------------------------------------------------------
    # Helper: find best matching column
    # -------------------------------------------------------
    def _get_best_col(df, config_key, fallbacks):

        config_val = b.get(config_key, [])
        search_terms = []

        if isinstance(config_val, list):
            search_terms.extend(config_val)

        elif isinstance(config_val, str):
            search_terms.append(config_val)

        search_terms.extend(fallbacks)

        for term in search_terms:

            if not term:
                continue

            # Exact match
            if term in df.columns:
                return term

            # Case-insensitive match
            for actual_col in df.columns:

                if (
                    str(actual_col).strip().lower()
                    == str(term).strip().lower()
                ):
                    return actual_col

        return None

    # -------------------------------------------------------
    # Detect columns
    # -------------------------------------------------------

    col_program_type = _get_best_col(
        df,
        "type_of_program",
        ["Program Type", "Type of Program", "Status", "Live/Repeat"]
    )

    col_market = _get_best_col(
        df,
        "market",
        ["Market", "Region", "Country"]
    )

    col_broadcaster = _get_best_col(
        df,
        "broadcaster",
        ["Broadcaster", "Station"]
    )

    col_channel = _get_best_col(
        df,
        "tv_channel",
        ["Channel", "TV Channel", "Channel ID"]
    )

    col_date = _get_best_col(
        df,
        "date",
        ["Date", "BSR_Local_Date", "BSR Date"]
    )

    col_start = _get_best_col(
        df,
        "start_time",
        ["Start", "Program Start", "Start Time"]
    )

    col_end = _get_best_col(
        df,
        "end_time",
        ["End", "End Time"]
    )

    # IMPORTANT:
    # Use the strongest match identifier possible
    col_match = _get_best_col(
        df,
        "phase_fixture_episode",
        [
            "PhaseFixtureEpisode",
            "Program title",
            "Combined (translated)"
        ]
    )

    # -------------------------------------------------------
    # Validate required columns
    # -------------------------------------------------------

    required_cols_map = {
        "Program Type": col_program_type,
        "Market": col_market,
        "Broadcaster": col_broadcaster,
        "Channel": col_channel,
        "Date": col_date,
        "Start": col_start,
        "End": col_end,
        "Match Identifier": col_match
    }

    missing_headers = [
        label
        for label, val in required_cols_map.items()
        if not val
    ]

    if missing_headers:

        df[result_col] = (
            f"Error: Missing Headers ({', '.join(missing_headers)})"
        )

        return df

    # -------------------------------------------------------
    # LIVE rows only
    # -------------------------------------------------------

    live_mask = (
        df[col_program_type]
        .astype(str)
        .str.strip()
        .str.lower()
        .eq("live")
    )

    # No live rows
    if not live_mask.any():

        df[result_col] = "Not Applicable"

        return df

    # -------------------------------------------------------
    # Normalize fields
    # -------------------------------------------------------

    for col in [
        col_market,
        col_broadcaster,
        col_channel,
        col_date,
        col_start,
        col_end,
        col_match
    ]:

        df[col] = (
            df[col]
            .astype(str)
            .str.strip()
            .str.lower()
        )

    # -------------------------------------------------------
    # Duplicate detection logic
    # -------------------------------------------------------

    # Exact LIVE duplicate definition:
    # same market + broadcaster + channel
    # same date + start + end
    # same match/program

    group_cols = [
        col_market,
        col_broadcaster,
        col_channel,
        col_date,
        col_start,
        col_end,
        col_match
    ]

    live_df = df[live_mask]

    duplicate_mask = live_df.duplicated(
        subset=group_cols,
        keep="first"
    )

    # -------------------------------------------------------
    # Assign results
    # -------------------------------------------------------

    # Default for live rows
    df.loc[live_mask, result_col] = "True"

    # Only repeated duplicate rows become False
    df.loc[
        duplicate_mask[duplicate_mask].index,
        result_col
    ] = "False"

    # Non-live rows
    df.loc[~live_mask, result_col] = "Not Applicable"

    return df
# -----------------------------------------------------------
# 16️⃣ Metered Channel Estimation Check (Robust Column Matching)
# -----------------------------------------------------------
def metered_channel_estimation_check(df, bsr_cols):
    """
    Validates metered channels using Market + Channel ID combination.
    Skips channels where Source in master list = 'Broadcaster Data'
    """

    df = df.copy()
    df["Metered_Estimation_Check_OK"] = True
    df["Metered_Estimation_Check_Remark"] = "OK"

    # --- 1. Load Local Master List ---
    master_list_path = "master_metered_list.xlsx"

    if not os.path.exists(master_list_path):
        df["Metered_Estimation_Check_OK"] = False
        df["Metered_Estimation_Check_Remark"] = "Error: master_metered_list.xlsx not found."
        return df

    try:
        metered_list_df = pd.read_excel(master_list_path)

        # Detect columns
        m_col_market = _find_column(metered_list_df, ["market"])
        m_col_ch_id = _find_column(metered_list_df, ["channel id", "channel_id"])
        m_col_source = _find_column(metered_list_df, ["source"])

        if not m_col_market or not m_col_ch_id:
            df["Metered_Estimation_Check_OK"] = False
            df["Metered_Estimation_Check_Remark"] = (
                f"Error: Master list headers mismatch. Found: {list(metered_list_df.columns)}"
            )
            return df

        # --- Create reference sets ---
        metered_reference_set = set()
        broadcaster_skip_set = set()

        for _, row in metered_list_df.iterrows():
            market_val = str(row.get(m_col_market, "")).strip().lower()
            channel_id_val = str(row.get(m_col_ch_id, "")).strip().lower()
            key = (market_val, channel_id_val)

            source_val = str(row.get(m_col_source, "")).strip().lower() if m_col_source else ""

            if source_val == "broadcaster data":
                broadcaster_skip_set.add(key)
            else:
                metered_reference_set.add(key)

    except Exception as e:
        df["Metered_Estimation_Check_OK"] = False
        df["Metered_Estimation_Check_Remark"] = f"Error reading Master List: {e}"
        return df

    # --- 2. Resolve BSR Columns ---
    col_market = _find_column(df, bsr_cols.get("market"))
    col_ch_id = _find_column(df, bsr_cols.get("channel_id"))
    col_est_aud = _find_column(df, bsr_cols.get("aud_estimates"))
    col_met_aud = _find_column(df, bsr_cols.get("aud_metered"))

    # --- 3. Run Validation ---
    for idx, row in df.iterrows():
        market_val = str(row.get(col_market, "")).strip().lower()
        channel_id_val = str(row.get(col_ch_id, "")).strip().lower()

        key = (market_val, channel_id_val)

        # 🚫 Skip Broadcaster Data channels
        if key in broadcaster_skip_set:
            df.at[idx, "Metered_Estimation_Check_OK"] = True
            df.at[idx, "Metered_Estimation_Check_Remark"] = (
                "Skipped: Source is Broadcaster Data"
            )
            continue

        is_metered = key in metered_reference_set

        if is_metered:
            est_present = _is_present(row.get(col_est_aud))
            met_present = _is_present(row.get(col_met_aud))

            if est_present:
                df.at[idx, "Metered_Estimation_Check_OK"] = False
                df.at[idx, "Metered_Estimation_Check_Remark"] = (
                    f"Violation: Metered channel (Market: {row.get(col_market)}, "
                    f"Channel ID: {row.get(col_ch_id)}) has ESTIMATED data."
                )
            elif not met_present:
                df.at[idx, "Metered_Estimation_Check_OK"] = False
                df.at[idx, "Metered_Estimation_Check_Remark"] = (
                    f"Violation: Metered channel (Market: {row.get(col_market)}, "
                    f"Channel ID: {row.get(col_ch_id)}) is missing metered audience."
                )
        else:
            df.at[idx, "Metered_Estimation_Check_Remark"] = "Non-metered channel"

    return df

# -----------------------------------------------------------
# ✅ Excel Coloring for True/False checks
def color_excel(output_path, df):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    wb = load_workbook(output_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {name: idx + 1 for idx, name in enumerate(headers)}

    # ✅ Support both old and new QC result columns
    qc_columns = [
        col for col in df.columns
        if col.endswith("_OK") or col.endswith("_result")
    ]

    for col_name in qc_columns:
        if col_name in col_map:
            col_idx = col_map[col_name]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                if val in [True, "True"]:
                    cell.fill = GREEN_FILL
                elif val in [False, "False"]:
                    cell.fill = RED_FILL
                elif val in ["NA", None]:
                    cell.fill = GREY_FILL

    wb.save(output_path)
# -----------------------------------------------------------
# Summary Sheet
def generate_summary_sheet(output_path, df):
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = load_workbook(output_path)

    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws = wb.create_sheet("Summary")

    # ✅ Support both old and new QC result columns
    qc_columns = [
        col for col in df.columns
        if col.endswith("_OK") or col.endswith("_result")
    ]

    summary_rows = []

    for col in qc_columns:
        series = df[col].astype(str)

        passed = series.eq("True").sum()
        failed = series.eq("False").sum()
        na = series.eq("NA").sum()
        total = passed + failed + na

        summary_rows.append([
            col,
            int(total),
            int(passed),
            int(failed),
            int(na)
        ])

    summary_df = pd.DataFrame(
        summary_rows,
        columns=["Check", "Total Evaluated", "Passed", "Failed", "NA"]
    )

    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)

    wb.save(output_path)